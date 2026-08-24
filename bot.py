"""
💰 HISOB-KITOB BOTI  v3

- Bir nechta biznes, har yozuvda bot qaysi biznes ekanini so'raydi
- Kategoriyalar yo'q: faqat KIRIM va CHIQIM
- Ikkita kassa: 💵 Naqd va 💳 Karta
- Kassa qoldig'i avtomatik yuritiladi, tekshirish va pul ko'chirish bor

    pip install aiogram openpyxl
    python bot.py
"""
from __future__ import annotations

import asyncio
import datetime as dt
import io
import logging
import os
import re
import sqlite3
import sys
import shutil
from collections import defaultdict
from pathlib import Path

try:
    from aiogram import Bot, Dispatcher, F
    from aiogram.client.default import DefaultBotProperties
    from aiogram.enums import ParseMode
    from aiogram.filters import Command, CommandStart, StateFilter
    from aiogram.fsm.context import FSMContext
    from aiogram.fsm.state import State, StatesGroup
    from aiogram.fsm.storage.memory import MemoryStorage
    from aiogram.types import (
        BufferedInputFile, CallbackQuery, InlineKeyboardButton,
        InlineKeyboardMarkup, KeyboardButton, Message, ReplyKeyboardMarkup, WebAppInfo,
    )
except ImportError:
    sys.exit("Avval kutubxonalarni o'rnating:\n\n    pip install aiogram openpyxl\n")

import webapp                                    # Mini App serveri (standart kutubxona)

BASE = Path(__file__).resolve().parent

# Bulutda ishlayapmizmi? (Railway, Render va h.k. PORT o'zgaruvchisini beradi)
BULUT = bool(os.getenv("PORT") or os.getenv("RAILWAY_PUBLIC_DOMAIN")
             or os.getenv("RENDER_EXTERNAL_HOSTNAME"))


def _data_dir() -> Path:
    """Baza saqlanadigan papka.

    Bulutli xizmatlarda dastur papkasi har yangilanishda tozalanadi.
    Shuning uchun doimiy disk (volume) ulanadi — odatda /data.
    Shunday papka bo'lsa va unga yozib bo'lsa, baza o'sha yerda turadi.
    """
    env = os.getenv("DATA_DIR", "").strip()
    if env:
        return Path(env)
    vol = Path("/data")
    try:
        if vol.is_dir():
            probe = vol / ".yozib_koraman"
            probe.write_text("1")
            probe.unlink()
            return vol
    except Exception:
        pass
    return BASE


DATA_DIR = _data_dir()
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "hisobot.db"

# Birinchi ishga tushish: loyiha bilan kelgan bazani doimiy diskka ko'chiramiz.
# Keyingi safar DB_PATH allaqachon mavjud bo'ladi va bu bo'lim o'tkazib yuboriladi —
# ya'ni serverdagi yozuvlar hech qachon eski nusxa bilan almashmaydi.
_boshlangich = BASE / "hisobot.db"
if DB_PATH != _boshlangich and not DB_PATH.exists() and _boshlangich.exists():
    import shutil as _sh
    _sh.copy2(_boshlangich, DB_PATH)
    print(f"[baza] Boshlang'ich nusxa ko'chirildi -> {DB_PATH}")

# Bulutli xizmatlar portni PORT orqali beradi
WEB_PORT = int(os.getenv("PORT") or os.getenv("WEB_PORT") or 8080)
# Bulutda: tashqariga ochiq + imzo tekshiruvi majburiy. Kompyuterda: aksincha.
WEB_DEV = os.getenv("MINIAPP_DEV", "0" if BULUT else "1") not in ("0", "false", "no")
WEB_HOST = os.getenv("MINIAPP_HOST", "0.0.0.0" if BULUT else "127.0.0.1")
ENV_PATH = BASE / "bot.env"

_TZ_NOMI = os.getenv("TIMEZONE", "Asia/Tashkent")
try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo(_TZ_NOMI)
except Exception:
    # Windows'da vaqt mintaqalari bazasi bo'lmasligi mumkin.
    # Toshkent doimiy UTC+5 (yozgi vaqt yo'q) — shuni ishlatamiz.
    try:
        _soat = int(os.getenv("UTC_OFFSET", "5"))
    except ValueError:
        _soat = 5
    TZ = dt.timezone(dt.timedelta(hours=_soat))


def hozir() -> dt.datetime:
    """Toshkent vaqti (server qayerda bo'lsa ham)."""
    return dt.datetime.now(TZ)


def bugun() -> dt.date:
    return hozir().date()


logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s")
log = logging.getLogger("hisobot")

# Ikkita kassa. Uchinchisi kerak bo'lsa shu ro'yxatga qo'shiladi.
KASSALAR = ["Naqd", "Karta"]
KASSA_EMOJI = {"Naqd": "💵", "Karta": "💳"}
# To'lov turlari: kassaga tushadigan ikkitasi + kassaga tegmaydigan "Qarzga"
TOLOVLAR = ["Naqd", "Karta", "Qarzga"]


# ═══════════════════════════════════════════════════════ 1. TOKEN
def get_token() -> str:
    token = os.getenv("BOT_TOKEN", "").strip()
    if not token and ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if line.startswith("BOT_TOKEN="):
                token = line.split("=", 1)[1].strip()
    if not token:
        print("\n" + "═" * 58)
        print("  💰 HISOB-KITOB BOTI — birinchi sozlash")
        print("═" * 58)
        print("\n  1. Telegramda @BotFather ni oching")
        print("  2. /newbot yuboring, botga nom bering")
        print("  3. U bergan tokenni shu yerga tashlang\n")
        try:
            token = input("  Token: ").strip()
        except (EOFError, KeyboardInterrupt):
            sys.exit("\nBekor qilindi.")
        if not token:
            sys.exit("Token kiritilmadi.")
        ENV_PATH.write_text(f"BOT_TOKEN={token}\n", encoding="utf-8")
        print(f"\n  ✅ Saqlandi: {ENV_PATH.name}\n")
    return token


# ═══════════════════════════════════════════════════════ 2. BAZA
_wal_qilindi = False


def db() -> sqlite3.Connection:
    """Har chaqiruvda yangi ulanish — bot va Mini App bir vaqtda ishlay oladi."""
    global _wal_qilindi
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    if not _wal_qilindi:
        try:
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            _wal_qilindi = True
        except sqlite3.Error:
            pass
    conn.execute("PRAGMA busy_timeout=15000")
    return conn


def init_db() -> None:
    with db() as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS bizneslar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nomi TEXT NOT NULL,
                emoji TEXT DEFAULT '🏢',
                faol INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS amaliyotlar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                biznes_id INTEGER NOT NULL DEFAULT 1,
                sana TEXT NOT NULL,
                vaqt TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                user_nomi TEXT NOT NULL,
                tur TEXT NOT NULL,
                izoh TEXT DEFAULT '',
                tolov TEXT DEFAULT 'Naqd',
                valyuta TEXT DEFAULT 'UZS',
                summa REAL NOT NULL,
                summa_uzs REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS qarzlar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                biznes_id INTEGER NOT NULL DEFAULT 1,
                sana TEXT NOT NULL,
                user_nomi TEXT NOT NULL,
                kim TEXT NOT NULL,
                turi TEXT NOT NULL,
                valyuta TEXT DEFAULT 'UZS',
                summa REAL NOT NULL,
                summa_uzs REAL NOT NULL,
                qaytarilgan REAL DEFAULT 0,
                holati TEXT DEFAULT 'Ochiq',
                izoh TEXT DEFAULT '',
                muddat TEXT
            );

            CREATE TABLE IF NOT EXISTS foydalanuvchilar (
                user_id INTEGER PRIMARY KEY,
                nomi TEXT,
                rol TEXT DEFAULT 'xodim',
                biznes_id INTEGER,
                qoshilgan TEXT
            );

            CREATE TABLE IF NOT EXISTS biznes_kalit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                biznes_id INTEGER NOT NULL,
                soz TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS kassa (
                biznes_id INTEGER PRIMARY KEY,
                naqd REAL DEFAULT 0,
                karta REAL DEFAULT 0,
                sozlangan INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS kochirish (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sana TEXT NOT NULL,
                vaqt TEXT NOT NULL,
                user_nomi TEXT NOT NULL,
                biznes_id INTEGER NOT NULL,
                qayerdan TEXT NOT NULL,
                biznes_2 INTEGER,
                qayerga TEXT NOT NULL,
                summa REAL NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sozlamalar (
                kalit TEXT PRIMARY KEY,
                qiymat TEXT
            );
            """
        )
        c.execute("INSERT OR IGNORE INTO sozlamalar (kalit,qiymat) VALUES ('usd_kurs','12800')")

        # eski bazadan yangilash
        for table, column, decl in (
            ("amaliyotlar", "biznes_id", "INTEGER NOT NULL DEFAULT 1"),
            ("qarzlar", "biznes_id", "INTEGER NOT NULL DEFAULT 1"),
            ("foydalanuvchilar", "biznes_id", "INTEGER"),
            ("qarzlar", "muddat", "TEXT"),
        ):
            cols = [r["name"] for r in c.execute(f"PRAGMA table_info({table})")]
            if column not in cols:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")

        cols = [r["name"] for r in c.execute("PRAGMA table_info(amaliyotlar)")]
        if "kategoriya" in cols:
            # kategoriyalar ishlatilmaydi — eski qiymatlar tozalanadi
            c.execute("UPDATE amaliyotlar SET kategoriya=''")
        # eski to'lov turlarini ikkita kassaga moslash
        c.execute("UPDATE amaliyotlar SET tolov='Karta' "
                  "WHERE tolov IN ('Plastik karta','Click/Payme','Bank o''tkazma')")

        bor_biznes = c.execute("SELECT COUNT(*) n FROM bizneslar").fetchone()["n"]
        bor_yozuv = c.execute("SELECT COUNT(*) n FROM amaliyotlar").fetchone()["n"]
        if not bor_biznes and bor_yozuv:
            c.execute("INSERT INTO bizneslar (id,nomi,emoji) VALUES (1,'Asosiy biznes','🏪')")
        c.execute("CREATE INDEX IF NOT EXISTS idx_sana ON amaliyotlar(sana)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_biz ON amaliyotlar(biznes_id)")


def cfg_get(key: str, default: str = "") -> str:
    with db() as c:
        row = c.execute("SELECT qiymat FROM sozlamalar WHERE kalit=?", (key,)).fetchone()
    return row["qiymat"] if row else default


def cfg_set(key: str, value) -> None:
    with db() as c:
        c.execute("INSERT INTO sozlamalar (kalit,qiymat) VALUES (?,?) "
                  "ON CONFLICT(kalit) DO UPDATE SET qiymat=excluded.qiymat", (key, str(value)))


def rate() -> float:
    try:
        k = float(cfg_get("usd_kurs", "12800"))
    except ValueError:
        return 12800.0
    return k if k > 0 else 12800.0


# ---- bizneslar -----------------------------------------------------------
EMOJIS = ["🏪", "🏭", "🚚", "🍽", "💈", "🏗", "💻", "🚗"]


def biz_all() -> list[sqlite3.Row]:
    with db() as c:
        return c.execute("SELECT * FROM bizneslar WHERE faol=1 ORDER BY id").fetchall()


def biz_get(biz_id: int) -> sqlite3.Row | None:
    with db() as c:
        return c.execute("SELECT * FROM bizneslar WHERE id=?", (biz_id,)).fetchone()


def biz_name(biz_id: int) -> str:
    b = biz_get(biz_id)
    return f"{b['emoji']} {b['nomi']}" if b else "—"


def biz_add(nomi: str) -> int:
    n = len(biz_all())
    with db() as c:
        cur = c.execute("INSERT INTO bizneslar (nomi,emoji) VALUES (?,?)",
                        (nomi, EMOJIS[n % len(EMOJIS)]))
        return cur.lastrowid


def biz_rename(biz_id: int, nomi: str) -> None:
    with db() as c:
        c.execute("UPDATE bizneslar SET nomi=? WHERE id=?", (nomi, biz_id))


def biz_set_emoji(biz_id: int, emoji: str) -> None:
    with db() as c:
        c.execute("UPDATE bizneslar SET emoji=? WHERE id=?", (emoji, biz_id))


def biz_hide(biz_id: int) -> None:
    with db() as c:
        c.execute("UPDATE bizneslar SET faol=0 WHERE id=?", (biz_id,))


def kalit_all(biz_id: int | None = None) -> list[sqlite3.Row]:
    with db() as c:
        if biz_id:
            return c.execute("SELECT * FROM biznes_kalit WHERE biznes_id=? ORDER BY soz",
                             (biz_id,)).fetchall()
        return c.execute("SELECT * FROM biznes_kalit ORDER BY biznes_id, soz").fetchall()


def kalit_add(biz_id: int, sozlar: str) -> int:
    yangi = [w.strip().lower() for w in sozlar.replace(";", ",").split(",") if w.strip()]
    bor = {r["soz"] for r in kalit_all(biz_id)}
    qoshildi = 0
    with db() as c:
        for w in yangi:
            if w not in bor:
                c.execute("INSERT INTO biznes_kalit (biznes_id,soz) VALUES (?,?)", (biz_id, w))
                qoshildi += 1
    return qoshildi


def kalit_clear(biz_id: int) -> None:
    with db() as c:
        c.execute("DELETE FROM biznes_kalit WHERE biznes_id=?", (biz_id,))


def kalitdan_biznes(matn: str, ruxsat: list[int]) -> int | None:
    """Izohdagi kalit so'zga qarab biznesni topadi (faqat bittasiga to'g'ri kelsa)."""
    past = " " + matn.lower() + " "
    topilgan = set()
    for r in kalit_all():
        if r["biznes_id"] in ruxsat and f" {r['soz']} " in past:
            topilgan.add(r["biznes_id"])
    return topilgan.pop() if len(topilgan) == 1 else None


# ---- foydalanuvchilar ----------------------------------------------------
def users_all() -> list[sqlite3.Row]:
    with db() as c:
        return c.execute("SELECT * FROM foydalanuvchilar ORDER BY rol, nomi").fetchall()


def user_get(uid: int) -> sqlite3.Row | None:
    with db() as c:
        return c.execute("SELECT * FROM foydalanuvchilar WHERE user_id=?", (uid,)).fetchone()


def user_add(uid: int, nomi: str, rol: str = "xodim", biznes_id: int | None = None) -> None:
    with db() as c:
        c.execute(
            "INSERT INTO foydalanuvchilar (user_id,nomi,rol,biznes_id,qoshilgan) "
            "VALUES (?,?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET "
            "nomi=excluded.nomi, rol=excluded.rol, biznes_id=excluded.biznes_id",
            (uid, nomi, rol, biznes_id, bugun().isoformat()),
        )


def user_touch(uid: int, nomi: str) -> None:
    with db() as c:
        c.execute("UPDATE foydalanuvchilar SET nomi=? WHERE user_id=?", (nomi, uid))


def user_remove(uid: int) -> bool:
    with db() as c:
        return c.execute("DELETE FROM foydalanuvchilar WHERE user_id=? AND rol!='admin'",
                         (uid,)).rowcount > 0


def has_admin() -> bool:
    with db() as c:
        return c.execute("SELECT 1 FROM foydalanuvchilar WHERE rol='admin'").fetchone() is not None


def is_admin(uid: int) -> bool:
    u = user_get(uid)
    return bool(u and u["rol"] == "admin")


def allowed(uid: int) -> bool:
    return user_get(uid) is not None


def my_businesses(uid: int) -> list[sqlite3.Row]:
    u = user_get(uid)
    if not u:
        return []
    if u["rol"] == "admin" or u["biznes_id"] is None:
        return biz_all()
    b = biz_get(u["biznes_id"])
    return [b] if b else []


# ---- amaliyotlar ---------------------------------------------------------
def tx_add(**k) -> int:
    with db() as c:
        cur = c.execute(
            "INSERT INTO amaliyotlar (biznes_id,sana,vaqt,user_id,user_nomi,tur,izoh,"
            "tolov,valyuta,summa,summa_uzs) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (k["biznes_id"], k["sana"], k["vaqt"], k["user_id"], k["user_nomi"], k["tur"],
             k["izoh"], k["tolov"], k["valyuta"], k["summa"], k["summa_uzs"]),
        )
        return cur.lastrowid


def tx_delete(tx_id: int, biz_ids: list[int]) -> bool:
    if not biz_ids:
        return False
    marks = ",".join("?" * len(biz_ids))
    with db() as c:
        return c.execute(f"DELETE FROM amaliyotlar WHERE id=? AND biznes_id IN ({marks})",
                         (tx_id, *biz_ids)).rowcount > 0


def tx_period(start: str, end: str, biz_id: int | None = None) -> list[sqlite3.Row]:
    q = "SELECT * FROM amaliyotlar WHERE sana BETWEEN ? AND ?"
    args: list = [start, end]
    if biz_id:
        q += " AND biznes_id=?"
        args.append(biz_id)
    with db() as c:
        return c.execute(q + " ORDER BY sana, vaqt", args).fetchall()


def tx_last(n: int, biz_ids: list[int]) -> list[sqlite3.Row]:
    if not biz_ids:
        return []
    marks = ",".join("?" * len(biz_ids))
    with db() as c:
        return c.execute(
            f"SELECT * FROM amaliyotlar WHERE biznes_id IN ({marks}) ORDER BY id DESC LIMIT ?",
            (*biz_ids, n)).fetchall()


def tx_all(biz_id: int | None = None) -> list[sqlite3.Row]:
    q = "SELECT * FROM amaliyotlar"
    args: tuple = ()
    if biz_id:
        q += " WHERE biznes_id=?"
        args = (biz_id,)
    with db() as c:
        return c.execute(q + " ORDER BY id", args).fetchall()


# ---- qarzlar -------------------------------------------------------------
def debt_add(**k) -> int:
    with db() as c:
        cur = c.execute(
            "INSERT INTO qarzlar (biznes_id,sana,user_nomi,kim,turi,valyuta,summa,"
            "summa_uzs,izoh,muddat) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (k["biznes_id"], k["sana"], k["user_nomi"], k["kim"], k["turi"],
             k["valyuta"], k["summa"], k["summa_uzs"], k["izoh"], k.get("muddat")))
        return cur.lastrowid


def debts_all(only_open: bool = False, biz_id: int | None = None) -> list[sqlite3.Row]:
    q = "SELECT * FROM qarzlar WHERE 1=1"
    args: list = []
    if only_open:
        q += " AND holati='Ochiq'"
    if biz_id:
        q += " AND biznes_id=?"
        args.append(biz_id)
    with db() as c:
        return c.execute(q + " ORDER BY id", args).fetchall()


def debt_get(debt_id: int, biz_ids: list[int]) -> sqlite3.Row | None:
    if not biz_ids:
        return None
    marks = ",".join("?" * len(biz_ids))
    with db() as c:
        return c.execute(f"SELECT * FROM qarzlar WHERE id=? AND biznes_id IN ({marks})",
                         (debt_id, *biz_ids)).fetchone()


def debt_close(debt_id: int) -> None:
    with db() as c:
        c.execute("UPDATE qarzlar SET holati='Yopilgan', qaytarilgan=summa_uzs WHERE id=?",
                  (debt_id,))


# ---- kassa ---------------------------------------------------------------
def kassa_row(biz_id: int) -> sqlite3.Row:
    with db() as c:
        row = c.execute("SELECT * FROM kassa WHERE biznes_id=?", (biz_id,)).fetchone()
        if row is None:
            c.execute("INSERT INTO kassa (biznes_id,naqd,karta,sozlangan) VALUES (?,0,0,0)",
                      (biz_id,))
            row = c.execute("SELECT * FROM kassa WHERE biznes_id=?", (biz_id,)).fetchone()
    return row


def kassa_sozlangan(biz_id: int) -> bool:
    return bool(kassa_row(biz_id)["sozlangan"])


def kassa_set_boshlangich(biz_id: int, naqd: float, karta: float) -> None:
    kassa_row(biz_id)
    with db() as c:
        c.execute("UPDATE kassa SET naqd=?, karta=?, sozlangan=1 WHERE biznes_id=?",
                  (naqd, karta, biz_id))


def kochirish_add(**k) -> int:
    with db() as c:
        cur = c.execute(
            "INSERT INTO kochirish (sana,vaqt,user_nomi,biznes_id,qayerdan,biznes_2,qayerga,summa)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (k["sana"], k["vaqt"], k["user_nomi"], k["biznes_id"], k["qayerdan"],
             k.get("biznes_2"), k["qayerga"], k["summa"]))
        return cur.lastrowid


def kochirish_all(biz_id: int | None = None) -> list[sqlite3.Row]:
    with db() as c:
        if biz_id:
            return c.execute("SELECT * FROM kochirish WHERE biznes_id=? OR biznes_2=? ORDER BY id",
                             (biz_id, biz_id)).fetchall()
        return c.execute("SELECT * FROM kochirish ORDER BY id").fetchall()


def kassa_holati(biz_id: int, kunga: str | None = None) -> dict[str, float]:
    """Biznesning naqd va karta qoldig'i. kunga berilsa — o'sha kun oxiriga."""
    k = kassa_row(biz_id)
    bal = {"Naqd": float(k["naqd"] or 0), "Karta": float(k["karta"] or 0)}
    with db() as c:
        q = "SELECT tur,tolov,SUM(summa_uzs) s FROM amaliyotlar WHERE biznes_id=?"
        args: list = [biz_id]
        if kunga:
            q += " AND sana<=?"
            args.append(kunga)
        for r in c.execute(q + " GROUP BY tur,tolov", args):
            if r["tolov"] not in bal:      # "Qarzga" kassaga tegmaydi
                continue
            bal[r["tolov"]] += r["s"] if r["tur"] == "Kirim" else -r["s"]

        q = "SELECT * FROM kochirish WHERE (biznes_id=? OR biznes_2=?)"
        args = [biz_id, biz_id]
        if kunga:
            q += " AND sana<=?"
            args.append(kunga)
        for r in c.execute(q, args):
            if r["biznes_id"] == biz_id and r["qayerdan"] in bal:
                bal[r["qayerdan"]] -= r["summa"]
            target_biz = r["biznes_2"] or r["biznes_id"]
            if target_biz == biz_id and r["qayerga"] in bal:
                bal[r["qayerga"]] += r["summa"]
    return bal


# ═══════════════════════════════════════════════════════ 3. YORDAMCHILAR
def fmt(n: float) -> str:
    return f"{round(n):,}".replace(",", " ")


AMOUNT_RE = re.compile(r"^\s*([\d\s.,]+)\s*(\$|usd|dollar|so'm|som|uzs)?\s*$", re.I)
GROUPED_RE = re.compile(r"^\d{1,3}([.,]\d{3})+$")     # 150.000 · 1,500,000 · 1.500.000


def parse_amount(text: str) -> tuple[float, str] | None:
    """Summani o'qiydi.

    Qoida (so'm bilan ishlashga moslangan):
      150000 · 150 000 · 150.000 · 150,000   -> 150 000
      1.500.000 · 1,500,000                  -> 1 500 000
      1,5 mln · 0.5 mln · 300 ming           -> ko'paytiruvchi bilan kasr ishlaydi
      20.50$ · 20,5$                         -> dollarda kasr ishlaydi
    Ya'ni ajratgichdan keyin ROSA 3 raqam kelsa — bu minglik ajratgich,
    1-2 raqam kelsa — kasr.
    """
    if not text:
        return None
    t = text.strip().lower()

    mult = 1
    for suffix, factor in (("mln", 1_000_000), ("million", 1_000_000),
                           ("ming", 1_000), ("k", 1_000)):
        if t.endswith(suffix):
            t, mult = t[: -len(suffix)].strip(), factor
            break

    m = AMOUNT_RE.match(t)
    if not m:
        soz = sozdan_raqam(t)          # «ikki yuz ellik ming»
        if soz:
            qiymat = soz * mult
            return (round(qiymat, 2), "UZS") if 0 < qiymat < 1e15 else None
        return None
    raw = m.group(1).replace(" ", "").replace("\u00a0", "")
    if not raw:
        return None

    if GROUPED_RE.match(raw):                 # 150.000 / 1,500,000 -> minglik
        raw = raw.replace(".", "").replace(",", "")
    else:
        oxirgi = max(raw.rfind("."), raw.rfind(","))
        if oxirgi == -1:
            pass                              # ajratgich yo'q
        else:
            butun = raw[:oxirgi].replace(".", "").replace(",", "")
            qism = raw[oxirgi + 1:]
            if len(qism) == 3 and mult == 1:
                raw = butun + qism            # 150.000 -> 150000
            else:
                raw = butun + "." + qism      # 1,5 -> 1.5   ·  20,50 -> 20.5

    try:
        value = float(raw) * mult
    except ValueError:
        return None
    if value <= 0 or value > 1e15:
        return None
    unit = (m.group(2) or "").lower()
    return round(value, 2), ("USD" if unit in ("$", "usd", "dollar") else "UZS")


def totals(rows) -> tuple[float, float]:
    inc = sum(r["summa_uzs"] for r in rows if r["tur"] == "Kirim")
    exp = sum(r["summa_uzs"] for r in rows if r["tur"] == "Chiqim")
    return inc, exp


def kassa_qatori(biz_id: int) -> str:
    bal = kassa_holati(biz_id)
    return (f"  💵 Naqd: <b>{fmt(bal['Naqd'])}</b> · "
            f"💳 Karta: <b>{fmt(bal['Karta'])}</b> so'm")


def _foiz(hozirgi: float, oldingi: float) -> str:
    """▲ +38%  ·  ▼ −22%  ·  bo'sh (taqqoslab bo'lmasa)."""
    if oldingi == 0:
        return " <i>(yangi)</i>" if hozirgi else ""
    farq = (hozirgi - oldingi) / abs(oldingi) * 100
    if abs(farq) < 1:
        return " ▬ <i>o'zgarishsiz</i>"
    belgi = "▲" if farq > 0 else "▼"
    return f" {belgi} <b>{farq:+.0f}%</b>"


def oldingi_davr(start: str, end: str) -> tuple[str, str]:
    """Shuncha uzunlikdagi oldingi davr."""
    s_d = dt.date.fromisoformat(start)
    e_d = dt.date.fromisoformat(end)
    uzunlik = (e_d - s_d).days + 1
    return ((s_d - dt.timedelta(days=uzunlik)).isoformat(),
            (e_d - dt.timedelta(days=uzunlik)).isoformat())


def report(start: str, end: str, title: str, biz_ids: list[int],
           kassa_bilan: bool = True, taqqoslash: bool = True) -> str:
    parts = [f"<b>{title}</b>", ""]
    grand: list = []
    o_start, o_end = oldingi_davr(start, end) if taqqoslash else ("", "")
    eski_jami = 0.0
    for b in biz_all():
        if b["id"] not in biz_ids:
            continue
        rows = tx_period(start, end, b["id"])
        grand += rows
        inc, exp = totals(rows)
        profit = inc - exp
        farq = ""
        if taqqoslash:
            e_inc, e_exp = totals(tx_period(o_start, o_end, b["id"]))
            eski_jami += e_inc - e_exp
            farq = _foiz(profit, e_inc - e_exp)
        parts += [
            f"<b>{b['emoji']} {b['nomi']}</b>",
            f"  🟢 Kirim:  {fmt(inc)}",
            f"  🔴 Chiqim: {fmt(exp)}",
            f"  {'💰' if profit >= 0 else '⚠️'} Foyda:  <b>{fmt(profit)}</b> so'm{farq}",
            f"  <i>{len(rows)} ta yozuv</i>",
        ]
        if kassa_bilan:
            parts.append(kassa_qatori(b["id"]))
        parts.append("")
    if len(biz_ids) > 1:
        inc, exp = totals(grand)
        profit = inc - exp
        parts += [
            "═════════════════",
            "<b>📌 UMUMIY</b>",
            f"  🟢 Kirim:  <b>{fmt(inc)}</b> so'm",
            f"  🔴 Chiqim: <b>{fmt(exp)}</b> so'm",
            f"  {'💰' if profit >= 0 else '⚠️'} Foyda:  <b>{fmt(profit)}</b> so'm  "
            f"(${profit / rate():,.2f})" + (_foiz(profit, eski_jami) if taqqoslash else ""),
        ]
        if kassa_bilan:
            naqd = sum(kassa_holati(i)["Naqd"] for i in biz_ids)
            karta = sum(kassa_holati(i)["Karta"] for i in biz_ids)
            parts.append(f"  💵 Naqd: <b>{fmt(naqd)}</b> · 💳 Karta: <b>{fmt(karta)}</b> so'm")
    return "\n".join(parts).strip()


def kassa_report(biz_ids: list[int]) -> str:
    lines = ["💰 <b>KASSA HOLATI</b>",
             f"<i>{hozir():%d.%m.%Y %H:%M}</i>", ""]
    jami = {"Naqd": 0.0, "Karta": 0.0}
    for b in biz_all():
        if b["id"] not in biz_ids:
            continue
        bal = kassa_holati(b["id"])
        for k in KASSALAR:
            jami[k] += bal[k]
        lines.append(f"<b>{b['emoji']} {b['nomi']}</b>")
        for k in KASSALAR:
            ogoh = "  ⚠️" if bal[k] < 0 else ""
            lines.append(f"  {KASSA_EMOJI[k]} {k}:  <b>{fmt(bal[k])}</b>{ogoh}")
        lines.append(f"  <i>Jami: {fmt(sum(bal.values()))} so'm</i>")
        if not kassa_sozlangan(b["id"]):
            lines.append("  <i>⚙️ boshlang'ich qoldiq hali kiritilmagan</i>")
        lines.append("")
    if len(biz_ids) > 1:
        t = jami["Naqd"] + jami["Karta"]
        lines += ["═════════════════",
                  f"📌 <b>UMUMIY: {fmt(t)} so'm</b>",
                  f"   💵 {fmt(jami['Naqd'])} · 💳 {fmt(jami['Karta'])}",
                  f"   <i>≈ ${t / rate():,.2f}</i>"]
    if any(kassa_holati(i)[k] < 0 for i in biz_ids for k in KASSALAR):
        lines += ["", "⚠️ <i>Kassa minusda — yozuv tushib qolgan bo'lishi mumkin.</i>"]
    return "\n".join(lines).strip()


def kassa_harakati(biz_id: int, kunlar: int = 7) -> str:
    bugun = bugun()
    start = bugun - dt.timedelta(days=kunlar - 1)
    b = biz_get(biz_id)
    lines = [f"📆 <b>{b['emoji']} {b['nomi']} — KASSA HARAKATI</b>",
             f"<i>oxirgi {kunlar} kun</i>", ""]
    for i in range(kunlar):
        kun = start + dt.timedelta(days=i)
        rows = tx_period(kun.isoformat(), kun.isoformat(), biz_id)
        koch = [k for k in kochirish_all(biz_id) if k["sana"] == kun.isoformat()]
        if not rows and not koch:
            continue
        inc, exp = totals([r for r in rows if r["tolov"] in KASSALAR])
        oxiri = kassa_holati(biz_id, kun.isoformat())
        qism = []
        if inc:
            qism.append(f"+{fmt(inc)}")
        if exp:
            qism.append(f"−{fmt(exp)}")
        if koch:
            qism.append(f"🔁 {len(koch)} ta")
        lines.append(f"<code>{kun:%d.%m}</code>  {' '.join(qism)}")
        lines.append(f"   💵 <b>{fmt(oxiri['Naqd'])}</b> · 💳 <b>{fmt(oxiri['Karta'])}</b>")
    bal = kassa_holati(biz_id)
    lines += ["", f"<b>Hozir: 💵 {fmt(bal['Naqd'])} · 💳 {fmt(bal['Karta'])} so'm</b>"]
    return "\n".join(lines)


def debts_report(biz_ids: list[int]) -> str:
    lines = ["💳 <b>OCHIQ QARZLAR</b>", ""]
    jami_ol, jami_ber, bor = 0.0, 0.0, False
    for b in biz_all():
        if b["id"] not in biz_ids:
            continue
        open_ = debts_all(only_open=True, biz_id=b["id"])
        if not open_:
            continue
        bor = True
        lines.append(f"<b>{b['emoji']} {b['nomi']}</b>")
        for title, turi, icon in (("Menga qarzdorlar", "Men qarz berdim", "🟢"),
                                  ("Men qarzdorman", "Men qarz oldim", "🔴")):
            group = [d for d in open_ if d["turi"] == turi]
            if not group:
                continue
            s = sum(d["summa_uzs"] - d["qaytarilgan"] for d in group)
            if turi == "Men qarz berdim":
                jami_ol += s
            else:
                jami_ber += s
            lines.append(f" {icon} {title} — {fmt(s)} so'm")
            for d in group:
                izoh = f" · {d['izoh']}" if d["izoh"] else ""
                lines.append(f"   <code>#{d['id']}</code> {d['kim']}: "
                             f"{fmt(d['summa_uzs'] - d['qaytarilgan'])} ({d['sana']}){izoh}")
        lines.append("")
    if not bor:
        return "💳 <b>Qarzlar</b>\n\nOchiq qarz yo'q 👍"
    if len(biz_ids) > 1:
        lines += ["═════════════════",
                  f"📌 <b>Umumiy:</b> olishim {fmt(jami_ol)} · berishim {fmt(jami_ber)} so'm",
                  f"    Farq: <b>{fmt(jami_ol - jami_ber)}</b> so'm"]
    return "\n".join(lines).strip()


def build_excel(biz_ids: list[int], kassa_bilan: bool = True) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(name="Arial", bold=True, color="FFFFFF")
    body = Font(name="Arial")
    money = '#,##0;(#,##0);-'
    wb = Workbook()
    wb.remove(wb.active)

    bizneslar = [b for b in biz_all() if b["id"] in biz_ids]
    ranges: list[tuple[str, str, int, int]] = []   # varaq, nomi, oxirgi qator, biznes_id

    for b in bizneslar:
        rows = tx_all(b["id"])
        safe = re.sub(r"[\\/*?:\[\]]", "-", b["nomi"])[:28] or f"Biznes {b['id']}"
        ws = wb.create_sheet(safe)
        heads = ["ID", "Sana", "Vaqt", "Kim yozdi", "Tur", "Izoh",
                 "To'lov", "Valyuta", "Summa", "Summa (so'm)"]
        for j, (h, w) in enumerate(zip(heads, [6, 12, 8, 18, 10, 34, 12, 9, 14, 16]), start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font, c.fill = hfont, hf
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        for i, r in enumerate(rows, start=2):
            for j, v in enumerate([r["id"], r["sana"], r["vaqt"], r["user_nomi"], r["tur"],
                                   r["izoh"], r["tolov"], r["valyuta"], r["summa"],
                                   r["summa_uzs"]], start=1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = body
                if j in (9, 10):
                    c.number_format = money
        ranges.append((safe, b["nomi"], max(len(rows) + 1, 2), b["id"]))

    sm = wb.create_sheet("Xulosa", 0)
    sm["A1"] = "UMUMIY XULOSA"
    sm["A1"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
    sm["A3"], sm["B3"] = "USD kursi", rate()
    heads = ["Biznes", "Kirim (so'm)", "Chiqim (so'm)", "Foyda (so'm)"]
    if kassa_bilan:
        heads += ["Kassa: naqd", "Kassa: karta"]
    for j, h in enumerate(heads, start=1):
        c = sm.cell(row=5, column=j, value=h)
        c.font, c.fill = hfont, hf
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        sm.column_dimensions[get_column_letter(j)].width = 20
    row = 6
    for safe, nomi, last, bid in ranges:
        q = "'" + safe.replace("'", "''") + "'"
        bal = kassa_holati(bid)
        sm.cell(row=row, column=1, value=nomi).font = Font(name="Arial", bold=True)
        sm.cell(row=row, column=2, value=f'=SUMIFS({q}!$J$2:$J${last},{q}!$E$2:$E${last},"Kirim")')
        sm.cell(row=row, column=3, value=f'=SUMIFS({q}!$J$2:$J${last},{q}!$E$2:$E${last},"Chiqim")')
        sm.cell(row=row, column=4, value=f"=B{row}-C{row}")
        if kassa_bilan:
            sm.cell(row=row, column=5, value=round(bal["Naqd"]))
            sm.cell(row=row, column=6, value=round(bal["Karta"]))
        for j in range(2, 7 if kassa_bilan else 5):
            c = sm.cell(row=row, column=j)
            c.font = body
            c.number_format = money
        row += 1
    if ranges:
        sm.cell(row=row, column=1, value="JAMI").font = Font(name="Arial", bold=True)
        for j in range(2, 7 if kassa_bilan else 5):
            col = get_column_letter(j)
            c = sm.cell(row=row, column=j, value=f"=SUM({col}6:{col}{row - 1})")
            c.font = Font(name="Arial", bold=True)
            c.number_format = money
            c.fill = PatternFill("solid", fgColor="EDEDED")
        if kassa_bilan:
            sm.cell(row=row + 2, column=1,
                    value="Kassa ustunlari bot hisobidan olingan (formula emas).").font = \
                Font(name="Arial", italic=True, size=9)

    # ---- Kunlik va oylik hisobot: barcha bizneslar bo'yicha ----
    def davr_varaq(nomi: str, kalit, sarlavha: str, tartib: list[str]) -> None:
        ws = wb.create_sheet(nomi)
        ws["A1"] = sarlavha
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
        ws.cell(row=3, column=1, value="Sana" if kalit == "kun" else "Oy")
        col = 2
        joylar: dict[int, int] = {}
        for b in bizneslar:
            ws.cell(row=2, column=col, value=b["nomi"])
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=col).font = Font(name="Arial", bold=True)
            for k, h in enumerate(["Kirim", "Chiqim", "Foyda"]):
                ws.cell(row=3, column=col + k, value=h)
            joylar[b["id"]] = col
            col += 3
        for k, h in enumerate(["Kirim", "Chiqim", "FOYDA"]):
            ws.cell(row=3, column=col + k, value=h)
        ws.cell(row=2, column=col, value="JAMI")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col).font = Font(name="Arial", bold=True)
        for j in range(1, col + 3):
            c = ws.cell(row=3, column=j)
            c.font, c.fill = hfont, hf
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(j)].width = 15
        ws.column_dimensions["A"].width = 13
        ws.freeze_panes = "B4"

        r = 4
        for key in tartib:
            ws.cell(row=r, column=1, value=key).font = Font(name="Arial", bold=True)
            for b in bizneslar:
                rows = [x for x in tx_all(b["id"])
                        if (x["sana"] if kalit == "kun" else x["sana"][:7]) == key]
                inc, exp = totals(rows)
                cc = joylar[b["id"]]
                for k, v in enumerate([inc, exp, inc - exp]):
                    c = ws.cell(row=r, column=cc + k, value=v)
                    c.number_format = money
                    c.font = body
            for k in range(3):
                letters = [get_column_letter(joylar[b["id"]] + k) for b in bizneslar]
                f = "=" + "+".join(f"{L}{r}" for L in letters) if letters else "=0"
                c = ws.cell(row=r, column=col + k, value=f)
                c.number_format = money
                c.font = Font(name="Arial", bold=(k == 2))
                if k == 2:
                    c.fill = PatternFill("solid", fgColor="EDEDED")
            r += 1
        if r > 4:
            ws.cell(row=r, column=1, value="JAMI").font = Font(name="Arial", bold=True)
            for j in range(2, col + 3):
                L = get_column_letter(j)
                c = ws.cell(row=r, column=j, value=f"=SUM({L}4:{L}{r - 1})")
                c.font = Font(name="Arial", bold=True)
                c.number_format = money
                c.fill = PatternFill("solid", fgColor="D9E2F3")

    kunlar, oylar = set(), set()
    for b in bizneslar:
        for x in tx_all(b["id"]):
            kunlar.add(x["sana"])
            oylar.add(x["sana"][:7])
    davr_varaq("Kunlik hisobot", "kun", "KUNLIK HISOBOT — barcha bizneslar",
               sorted(kunlar))
    davr_varaq("Oylik hisobot", "oy", "OYLIK HISOBOT — barcha bizneslar", sorted(oylar))

    dz = wb.create_sheet("Qarzlar")
    for j, h in enumerate(["ID", "Biznes", "Sana", "Kim yozdi", "Kim/Kimga", "Turi", "Valyuta",
                           "Summa", "Summa (so'm)", "Qaytarilgan", "Holati", "Izoh"], start=1):
        c = dz.cell(row=1, column=j, value=h)
        c.font, c.fill = hfont, hf
        dz.column_dimensions[get_column_letter(j)].width = 18
    i = 2
    for b in bizneslar:
        for d in debts_all(biz_id=b["id"]):
            for j, v in enumerate([d["id"], b["nomi"], d["sana"], d["user_nomi"], d["kim"],
                                   d["turi"], d["valyuta"], d["summa"], d["summa_uzs"],
                                   d["qaytarilgan"], d["holati"], d["izoh"]], start=1):
                c = dz.cell(row=i, column=j, value=v)
                c.font = body
                if j in (8, 9, 10):
                    c.number_format = money
            i += 1

    kz = wb.create_sheet("Pul ko'chirish")
    for j, h in enumerate(["ID", "Sana", "Vaqt", "Kim", "Biznes", "Qayerdan",
                           "Qayerga (biznes)", "Qayerga", "Summa"], start=1):
        c = kz.cell(row=1, column=j, value=h)
        c.font, c.fill = hfont, hf
        kz.column_dimensions[get_column_letter(j)].width = 18
    i = 2
    for k in kochirish_all():
        if k["biznes_id"] not in biz_ids:
            continue
        b2 = biz_get(k["biznes_2"])["nomi"] if k["biznes_2"] else ""
        for j, v in enumerate([k["id"], k["sana"], k["vaqt"], k["user_nomi"],
                               biz_get(k["biznes_id"])["nomi"], k["qayerdan"], b2,
                               k["qayerga"], k["summa"]], start=1):
            c = kz.cell(row=i, column=j, value=v)
            c.font = body
            if j == 9:
                c.number_format = money
        i += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════ 3b. MATNNI TUSHUNISH
SOZ_RAQAM = {
    "nol": 0, "bir": 1, "ikki": 2, "uch": 3, "tort": 4, "to'rt": 4, "besh": 5,
    "olti": 6, "yetti": 7, "sakkiz": 8, "to'qqiz": 9, "toqqiz": 9,
    "o'n": 10, "on": 10, "yigirma": 20, "o'ttiz": 30, "ottiz": 30,
    "qirq": 40, "ellik": 50, "oltmish": 60, "yetmish": 70,
    "sakson": 80, "to'qson": 90, "toqson": 90,
}
SOZ_KOPAYTMA = {
    "yuz": 100, "ming": 1_000, "million": 1_000_000, "mln": 1_000_000,
    "milion": 1_000_000, "millon": 1_000_000,
}
YARIM = {"yarim": 0.5, "yarm": 0.5}


def sozdan_raqam(matn: str) -> float | None:
    """«ikki yuz ellik ming» -> 250000.  Tushunmasa None."""
    sozlar = [w for w in re.split(r"[\s,]+", matn.lower().strip()) if w]
    if not sozlar:
        return None
    jami = 0.0
    joriy = 0.0
    topildi = False
    for w in sozlar:
        w = w.strip(".!?;:")
        if w in SOZ_RAQAM:
            joriy += SOZ_RAQAM[w]
            topildi = True
        elif w in YARIM:
            joriy += 0.5
            topildi = True
        elif w in SOZ_KOPAYTMA:
            k = SOZ_KOPAYTMA[w]
            if joriy == 0:
                joriy = 1
            if k == 100:
                joriy *= 100
            else:
                jami += joriy * k
                joriy = 0
            topildi = True
        elif w.replace(".", "").replace(",", "").isdigit():
            joriy += float(w.replace(",", "."))
            topildi = True
        else:
            return None            # notanish so'z — bu summa emas
    if not topildi:
        return None
    natija = jami + joriy
    return natija if natija > 0 else None


TOLOV_SOZLARI = [
    ("Karta", ("karta", "kartaga", "kartada", "plastik", "click", "klik", "payme",
               "paynet", "o'tkazma", "otkazma", "perechisleniye", "hisobga")),
    ("Qarzga", ("qarzga", "qarzdan", "nasiya", "keyin beradi", "keyin to'laydi",
                "keyin tolaydi")),
]


def tolovni_top(matn: str) -> str:
    past = " " + matn.lower() + " "
    for tolov, sozlar in TOLOV_SOZLARI:
        if any(f" {s} " in past or past.strip().endswith(" " + s) for s in sozlar):
            return tolov
    return "Naqd"


SANA_RE = re.compile(r"^(\d{1,2})[./](\d{1,2})(?:[./](\d{2,4}))?$")


def sanani_top(matn: str) -> tuple[dt.date | None, str]:
    """Matn boshidagi sana so'zini oladi: «kecha», «19.08», «3 kun oldin»."""
    t = matn.strip()
    past = t.lower()
    b = bugun()

    for soz, kun in (("bugun", 0), ("kecha", 1), ("kechagi", 1),
                     ("avvalgi kuni", 2), ("indinga", 2)):
        if past.startswith(soz):
            return b - dt.timedelta(days=kun), t[len(soz):].strip()

    m = re.match(r"^(\d{1,2})\s*kun\s*(oldin|avval)\b", past)
    if m:
        return b - dt.timedelta(days=int(m.group(1))), t[m.end():].strip()

    birinchi = past.split()[0] if past.split() else ""
    m = SANA_RE.match(birinchi)
    if m:
        kun, oy = int(m.group(1)), int(m.group(2))
        yil = int(m.group(3)) if m.group(3) else b.year
        if yil < 100:
            yil += 2000
        try:
            sana = dt.date(yil, oy, kun)
            if sana <= b:
                return sana, t[len(birinchi):].strip()
        except ValueError:
            pass
    return None, t



# ═══════════════════════════════════════════════════════ 4. KLAVIATURALAR
def main_menu(uid: int | None = None) -> ReplyKeyboardMarkup:
    """Kassa tugmasi faqat bot egasida bo'ladi."""
    rows = [[KeyboardButton(text="➕ Kirim"), KeyboardButton(text="➖ Chiqim")]]
    if uid is None or is_admin(uid):
        rows.append([KeyboardButton(text="💰 Kassa"), KeyboardButton(text="📊 Bugun")])
        rows.append([KeyboardButton(text="📅 Shu oy"), KeyboardButton(text="💳 Qarzlar")])
    else:
        rows.append([KeyboardButton(text="📊 Bugun"), KeyboardButton(text="📅 Shu oy")])
        rows.append([KeyboardButton(text="💳 Qarzlar")])
    rows.append([KeyboardButton(text="📥 Excel"), KeyboardButton(text="⚙️ Sozlamalar")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def grid(items: list[str], prefix: str, per_row: int = 2) -> InlineKeyboardMarkup:
    rows, row = [], []
    for i, item in enumerate(items):
        row.append(InlineKeyboardButton(text=item, callback_data=f"{prefix}:{i}"))
        if len(row) == per_row:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def tolov_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💵 Naqd", callback_data="pay:0"),
         InlineKeyboardButton(text="💳 Karta", callback_data="pay:1")],
        [InlineKeyboardButton(text="📝 Qarzga (kassaga tegmaydi)", callback_data="pay:2")],
        [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")],
    ])


def biz_kb(uid: int, prefix: str = "biz") -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"{b['emoji']} {b['nomi']}",
                                  callback_data=f"{prefix}:{b['id']}")]
            for b in my_businesses(uid)]
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def cancel_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")]])


def note_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⏭ Izohsiz davom etish", callback_data="note_skip")],
        [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")],
    ])


def kassa_kb(admin: bool) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text="📆 Kassa harakati", callback_data="kassa_hist")]]
    if admin:
        rows = [
            [InlineKeyboardButton(text="🧮 Kassani tekshirish", callback_data="kassa_check")],
            [InlineKeyboardButton(text="🔁 Pul ko'chirish", callback_data="kassa_move")],
            [InlineKeyboardButton(text="📆 Kassa harakati", callback_data="kassa_hist")],
            [InlineKeyboardButton(text="⚙️ Boshlang'ich qoldiq", callback_data="kassa_init")],
        ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def debts_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Ro'yxat", callback_data="debt_list")],
        [InlineKeyboardButton(text="➕ Men qarz berdim", callback_data="debt_new:out")],
        [InlineKeyboardButton(text="➕ Men qarz oldim", callback_data="debt_new:in")],
        [InlineKeyboardButton(text="✅ Qarzni yopish", callback_data="debt_close")],
    ])


def settings_kb(admin: bool) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text="🧾 Oxirgi 10 ta yozuv", callback_data="last10")]]
    if admin:
        rows += [
            [InlineKeyboardButton(text="🏢 Bizneslar", callback_data="biz_menu")],
            [InlineKeyboardButton(text="👥 Xodimlar", callback_data="users")],
            [InlineKeyboardButton(text="🔑 Kalit so'zlar", callback_data="kalitlar")],
            [InlineKeyboardButton(text="💱 USD kursi", callback_data="set_rate")],
            [InlineKeyboardButton(text="💾 Zaxira nusxa", callback_data="zaxira_ol")],
        ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def biz_menu_kb() -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"✏️ {b['emoji']} {b['nomi']}",
                                  callback_data=f"biz_ren:{b['id']}")] for b in biz_all()]
    rows.append([InlineKeyboardButton(text="➕ Yangi biznes qo'shish", callback_data="biz_new")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ═══════════════════════════════════════════════════════ 5. HOLATLAR
class Setup(StatesGroup):
    names = State()


class TxForm(StatesGroup):
    business = State()
    amount = State()
    note = State()
    payment = State()


class QuickBiz(StatesGroup):
    choose = State()


class EditForm(StatesGroup):
    summa = State()
    izoh = State()


class DebtForm(StatesGroup):
    business = State()
    who = State()
    amount = State()
    note = State()


class CloseDebt(StatesGroup):
    debt_id = State()
    kassaga = State()


class RateForm(StatesGroup):
    value = State()


class AddUser(StatesGroup):
    user_id = State()
    business = State()


class BizForm(StatesGroup):
    new_name = State()
    rename = State()


class KalitForm(StatesGroup):
    sozlar = State()


class KassaInit(StatesGroup):
    business = State()
    values = State()


class KassaCheck(StatesGroup):
    business = State()
    kassa = State()
    amount = State()


class KassaMove(StatesGroup):
    business = State()
    yonalish = State()
    target_biz = State()
    amount = State()


TOKEN = get_token()
bot = Bot(TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())


def uname(obj) -> str:
    u = obj.from_user
    return u.full_name or u.username or str(u.id)


def today_iso() -> str:
    return bugun().isoformat()


def my_ids(uid: int) -> list[int]:
    return [b["id"] for b in my_businesses(uid)]


def _biznes_yoq(uid: int) -> str:
    if not biz_all():
        if is_admin(uid):
            return ("🏢 Bizneslar ro'yxati bo'sh.\n\n"
                    "/start yuboring — bot nomlarini so'raydi.\n"
                    "Yoki ⚙️ Sozlamalar → 🏢 Bizneslar → ➕ dan qo'shing.")
        return "Bizneslar hali sozlanmagan. Bot egasiga ayting."
    return "Sizga biznes biriktirilmagan. Bot egasiga murojaat qiling."


# ═══════════════════════════════════════════════════════ 6. START
WELCOME = (
    "<b>⚡️ Eng tez usul — shunchaki yozing:</b>\n"
    "<code>+150000 savdo</code>\n"
    "<code>-45000 ijara</code>\n"
    "<code>+20$ xizmat</code>\n\n"
    "Plyus — kirim, minus — chiqim. Summadan keyingi so'zlar izoh bo'ladi.\n"
    "Men qaysi biznesga yozishni so'rayman — bitta tugma bosasiz.\n\n"
    "Pastdagi tugmalardan ham foydalanishingiz mumkin 👇"
)


@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    uid = message.from_user.id

    if not has_admin():
        user_add(uid, uname(message), "admin", None)
        await state.set_state(Setup.names)
        await message.answer(
            f"🎉 <b>Xush kelibsiz, {uname(message)}!</b>\n\n"
            "Siz bu botning <b>egasi</b>siz.\n\n"
            "Endi bizneslaringiz nomini yozing — <b>vergul bilan ajratib</b>, bitta xabarda:\n\n"
            "<code>Do'kon, Avtoservis</code>\n\n"
            "<i>Keyinchalik nomini o'zgartirish yoki yangisini qo'shish mumkin.</i>"
        )
        return

    if not allowed(uid):
        await message.answer(
            "⛔️ Sizda ruxsat yo'q.\n\n"
            f"Sizning ID: <code>{uid}</code>\n"
            "Shu raqamni bot egasiga yuboring — u sizni qo'shadi.")
        return

    user_touch(uid, uname(message))

    # bizneslar yo'qolib qolgan bo'lsa (yangi baza yoki eski v1 bazasi) — qayta so'raymiz
    if not biz_all():
        if is_admin(uid):
            await state.set_state(Setup.names)
            await message.answer(
                "🏢 <b>Bizneslar ro'yxati bo'sh.</b>\n\n"
                "Bu yangi baza bo'lsa yoki eski versiyadan o'tayotgan bo'lsangiz shunday bo'ladi.\n\n"
                "Bizneslaringiz nomini vergul bilan yozing:\n\n"
                "<code>Do'kon, Avtoservis</code>")
        else:
            await message.answer("Bizneslar hali sozlanmagan. Bot egasiga ayting.")
        return

    ro = "\n".join(f"  {b['emoji']} {b['nomi']}" for b in my_businesses(uid)) or "  —"
    await message.answer(
        f"Assalomu alaykum, <b>{uname(message)}</b>! 👋\n\n"
        f"<b>Bizneslaringiz:</b>\n{ro}\n\n" + WELCOME,
        reply_markup=main_menu(message.from_user.id))


MENYU_TUGMALARI = {"➕ Kirim", "➖ Chiqim", "💰 Kassa", "📊 Bugun", "📅 Shu oy",
                   "💳 Qarzlar", "📥 Excel", "⚙️ Sozlamalar"}


@dp.message(Setup.names, F.text)
async def setup_names(message: Message, state: FSMContext):
    if message.text.strip() in MENYU_TUGMALARI:
        await message.answer(
            "Bu menyu tugmasi — biznes nomi emas 🙂\n\n"
            "Bizneslaringiz nomini yozing, vergul bilan:\n"
            "<code>Do'kon, Avtoservis</code>")
        return
    names = [n.strip() for n in message.text.replace(";", ",").split(",") if n.strip()
             if n.strip() not in MENYU_TUGMALARI]
    if not names:
        await message.answer("Kamida bitta nom yozing. Masalan: <code>Do'kon, Avtoservis</code>")
        return
    await state.clear()
    for n in names[:8]:
        biz_add(n)
    ro = "\n".join(f"  {b['emoji']} {b['nomi']}" for b in biz_all())
    await message.answer(
        f"✅ <b>Bizneslar qo'shildi:</b>\n{ro}\n\n"
        "Endi <b>💰 Kassa → ⚙️ Boshlang'ich qoldiq</b> dan hozir kassangizda va "
        "kartangizda qancha pul borligini kiriting. Keyin bot kassani o'zi yuritadi.\n\n"
        + WELCOME,
        reply_markup=main_menu(message.from_user.id))


@dp.message(Command("id"))
async def cmd_id(message: Message):
    await message.answer(f"Sizning Telegram ID: <code>{message.from_user.id}</code>")


@dp.message(Command("yordam", "help"))
async def cmd_help(message: Message):
    await message.answer(
        "<b>📖 QO'LLANMA</b>\n\n"
        "<b>⚡️ Tez yozish</b>\n"
        "<code>+150000 savdo</code> — kirim\n"
        "<code>-45000 ijara</code> — chiqim\n"
        "<code>+20$ xizmat</code> — dollarda\n"
        "<code>-300 ming benzin</code> — «ming» va «mln» ishlaydi\n\n"
        "<b>💰 Kassa</b>\n"
        "Ikkita kassa: 💵 naqd va 💳 karta. Har yozuvdan keyin qoldiq ko'rsatiladi.\n"
        "«Qarzga» to'lovi kassaga tegmaydi — qarz yopilganda tushadi.\n"
        "🧮 Tekshirish — qo'ldagi pulni sanab, farqni yozib qo'yadi.\n"
        "🔁 Pul ko'chirish — naqddan kartaga yoki biznesdan biznesga.\n\n"
        "<b>⌨️ Buyruqlar</b>\n"
        "/kassa · /hafta · /yil · /excel · /ochirish 12 · /id\n"
        "/hisobot — kun yakunini hozir olish\n"
        "/shubhali — eski xato yozuvlarni topish\n"
        "/zaxira — bazaning nusxasini olish\n"
        "/qidir SO'Z — yozuvlarni qidirish\n"
        "/tahrir RAQAM — yozuvni tuzatish\n"
        "/kunlar — qaysi hafta kuni kuchli\n"
        "/xarajat — nimaga qancha ketgan\n"
        "/vaqt 22:00 — kunlik hisobot vaqtini o'zgartirish\n"
        "/ochir_xodim 123456789 (ega uchun)",
        reply_markup=main_menu(message.from_user.id))


# ═══════════════════════════════════════════════════════ 7. TEZ YOZISH
def qatorni_oqi(qator: str, ruxsat: list[int]) -> dict | None:
    """Bitta qatorni yozuvga aylantiradi.

    «kecha +450000 kartaga savdo» -> {sana, tur, summa, valyuta, tolov, izoh, biznes_id}
    """
    t = qator.strip()
    if not t:
        return None
    sana, t = sanani_top(t)
    t = t.strip()
    if not t or t[0] not in "+-":
        return None
    kind = "Kirim" if t[0] == "+" else "Chiqim"
    qolgan = t[1:].strip()
    if not qolgan:
        return None

    sozlar = qolgan.split()
    parsed, ishlatilgan = None, 0
    for take in (4, 3, 2, 1):                    # «ikki yuz ellik ming» ham
        if len(sozlar) >= take:
            nomzod = parse_amount(" ".join(sozlar[:take]))
            if nomzod:
                parsed, ishlatilgan = nomzod, take
                break
    if parsed is None:
        return None
    izoh = " ".join(sozlar[ishlatilgan:]).strip()
    return {
        "sana": sana or bugun(),
        "tur": kind,
        "summa": parsed[0],
        "valyuta": parsed[1],
        "tolov": tolovni_top(izoh),
        "izoh": izoh,
        "biznes_id": kalitdan_biznes(izoh, ruxsat),
    }


def yozuv_satri(y: dict) -> str:
    icon = "🟢" if y["tur"] == "Kirim" else "🔴"
    pul = f"{fmt(y['summa'])} $" if y["valyuta"] == "USD" else fmt(y["summa"])
    qism = f"{icon} <b>{pul}</b>"
    if y["tolov"] != "Naqd":
        qism += f" · {'💳' if y['tolov'] == 'Karta' else '📝'} {y['tolov']}"
    if y["izoh"]:
        qism += f" · {y['izoh']}"
    if y["biznes_id"]:
        b = biz_get(y["biznes_id"])
        qism += f"\n     {b['emoji']} {b['nomi']}" if b else ""
    if y["sana"] != bugun():
        qism += f"\n     📅 {y['sana']:%d.%m}"
    return qism


@dp.message(StateFilter(None), F.text.regexp(r"^\s*(?:[+\-]|\d{1,2}[./]\d{1,2}\s|kecha|bugun|\d+\s*kun)"))
async def quick_entry(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.clear()
    ruxsat = my_ids(message.from_user.id)
    if not ruxsat:
        await message.answer(_biznes_yoq(message.from_user.id))
        return

    qatorlar = [q for q in re.split(r"[\n;]+", message.text) if q.strip()]
    yozuvlar, xato = [], []
    for q in qatorlar:
        y = qatorni_oqi(q, ruxsat)
        (yozuvlar if y else xato).append(y or q.strip())
    if not yozuvlar:
        await message.answer(
            "Summani tushunmadim 🤔\n\n"
            "Masalan:\n<code>+150000 savdo</code>\n"
            "<code>-45000 kartaga ijara</code>\n"
            "<code>+ikki yuz ming xizmat</code>")
        return
    if xato:
        await message.answer("⚠️ Bu qator(lar)ni tushunmadim, o'tkazib yubordim:\n"
                             + "\n".join(f"• <code>{x}</code>" for x in xato[:5]))

    # biznesi aniqlanganlarni darhol saqlaymiz
    if len(ruxsat) == 1:
        for y in yozuvlar:
            y["biznes_id"] = ruxsat[0]
    if all(y["biznes_id"] for y in yozuvlar):
        await saqla_royxat(message, yozuvlar, message.from_user.id, uname(message))
        return

    await state.set_state(QuickBiz.choose)
    await state.update_data(yozuvlar=[{**y, "sana": y["sana"].isoformat()} for y in yozuvlar])
    sarlavha = ("📋 <b>Bitta yozuv:</b>" if len(yozuvlar) == 1
                else f"📋 <b>{len(yozuvlar)} ta yozuv tayyor:</b>")
    tugmalar = [[InlineKeyboardButton(text=f"{b['emoji']} {b['nomi']}",
                                      callback_data=f"qbiz:{b['id']}")]
                for b in my_businesses(message.from_user.id)]
    if len(yozuvlar) > 1:
        tugmalar.append([InlineKeyboardButton(text="🔀 Har biriga alohida",
                                              callback_data="qbiz:0")])
    tugmalar.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    await message.answer(
        sarlavha + "\n\n" + "\n".join(yozuv_satri(y) for y in yozuvlar) +
        "\n\n<b>Qaysi biznesga?</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=tugmalar))


async def saqla_royxat(target: Message, yozuvlar: list[dict], uid: int, author: str):
    """Bir nechta yozuvni saqlaydi va qisqacha xulosa beradi."""
    if len(yozuvlar) == 1:
        y = yozuvlar[0]
        await save_tx(target, biznes_id=y["biznes_id"], kind=y["tur"], amount=y["summa"],
                      currency=y["valyuta"], note=y["izoh"], payment=y["tolov"],
                      uid=uid, author=author, sana=y["sana"])
        return

    idlar, k_jami, c_jami = [], 0.0, 0.0
    for y in yozuvlar:
        uzs = y["summa"] * rate() if y["valyuta"] == "USD" else y["summa"]
        tx_id = tx_add(biznes_id=y["biznes_id"], sana=y["sana"].isoformat(),
                       vaqt=hozir().strftime("%H:%M"), user_id=uid, user_nomi=author,
                       tur=y["tur"], izoh=y["izoh"], tolov=y["tolov"],
                       valyuta=y["valyuta"], summa=round(y["summa"], 2), summa_uzs=round(uzs))
        idlar.append(tx_id)
        if y["tur"] == "Kirim":
            k_jami += uzs
        else:
            c_jami += uzs

    qatorlar = []
    for y, tx_id in zip(yozuvlar, idlar):
        b = biz_get(y["biznes_id"])
        icon = "🟢" if y["tur"] == "Kirim" else "🔴"
        uzs = y["summa"] * rate() if y["valyuta"] == "USD" else y["summa"]
        qatorlar.append(f"{icon} <code>#{tx_id}</code> {b['emoji'] if b else ''} "
                        f"<b>{fmt(uzs)}</b>" + (f" · {y['izoh']}" if y["izoh"] else ""))

    kassa_qism = ""
    if is_admin(uid):
        bizlar = {y["biznes_id"] for y in yozuvlar}
        kassa_qism = "\n" + "\n".join(
            f"💰 {biz_get(i)['nomi']}: 💵 {fmt(kassa_holati(i)['Naqd'])} · "
            f"💳 {fmt(kassa_holati(i)['Karta'])}" for i in sorted(bizlar))

    await target.answer(
        f"✅ <b>{len(idlar)} ta yozuv saqlandi</b>\n\n" + "\n".join(qatorlar) +
        f"\n\n🟢 Kirim: <b>{fmt(k_jami)}</b> · 🔴 Chiqim: <b>{fmt(c_jami)}</b>" + kassa_qism,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=f"↩️ Hammasini bekor qilish ({len(idlar)} ta)",
                                 callback_data="undo:" + ",".join(map(str, idlar)))]]))


@dp.callback_query(QuickBiz.choose, F.data.startswith("qbiz:"))
async def cb_quick_biz(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1])
    data = await state.get_data()
    yozuvlar = [{**y, "sana": dt.date.fromisoformat(y["sana"])} for y in data["yozuvlar"]]

    if biz_id == 0:                                   # har biriga alohida
        keyingi = next((i for i, y in enumerate(yozuvlar) if not y["biznes_id"]), None)
        if keyingi is None:
            await state.clear()
            await cb.message.edit_text("💾 Saqlandi")
            await saqla_royxat(cb.message, yozuvlar, cb.from_user.id, uname(cb))
            await cb.answer()
            return
        await state.update_data(alohida=True,
                                yozuvlar=[{**y, "sana": y["sana"].isoformat()} for y in yozuvlar])
        y = yozuvlar[keyingi]
        await cb.message.edit_text(
            f"{yozuv_satri(y)}\n\n<b>Qaysi biznesga?</b>",
            reply_markup=biz_kb(cb.from_user.id, "qone"))
        await cb.answer()
        return

    for y in yozuvlar:
        if not y["biznes_id"]:
            y["biznes_id"] = biz_id
    await state.clear()
    await cb.message.edit_text("💾 Saqlandi")
    await saqla_royxat(cb.message, yozuvlar, cb.from_user.id, uname(cb))
    await cb.answer()


@dp.callback_query(QuickBiz.choose, F.data.startswith("qone:"))
async def cb_quick_one(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1])
    data = await state.get_data()
    yozuvlar = [{**y, "sana": dt.date.fromisoformat(y["sana"])} for y in data["yozuvlar"]]
    for y in yozuvlar:
        if not y["biznes_id"]:
            y["biznes_id"] = biz_id
            break
    qolgan = [y for y in yozuvlar if not y["biznes_id"]]
    if qolgan:
        await state.update_data(yozuvlar=[{**y, "sana": y["sana"].isoformat()} for y in yozuvlar])
        await cb.message.edit_text(
            f"{yozuv_satri(qolgan[0])}\n\n<b>Qaysi biznesga?</b>",
            reply_markup=biz_kb(cb.from_user.id, "qone"))
    else:
        await state.clear()
        await cb.message.edit_text("💾 Saqlandi")
        await saqla_royxat(cb.message, yozuvlar, cb.from_user.id, uname(cb))
    await cb.answer()


# ═══════════════════════════════════════════════════════ 8. QADAMMA-QADAM
def tez_yozuvlar(biz_id: int, tur: str, n: int = 4) -> list[sqlite3.Row]:
    """Eng ko'p takrorlangan yozuvlar — tugma qilib chiqarish uchun."""
    chek = (bugun() - dt.timedelta(days=90)).isoformat()
    with db() as c:
        return c.execute(
            "SELECT izoh, tolov, valyuta, "
            "       CAST(ROUND(AVG(summa)) AS INTEGER) summa, COUNT(*) soni "
            "FROM amaliyotlar "
            "WHERE biznes_id=? AND tur=? AND sana>=? AND TRIM(COALESCE(izoh,''))<>'' "
            "GROUP BY LOWER(TRIM(izoh)), tolov, valyuta "
            "HAVING soni >= 2 ORDER BY soni DESC, summa DESC LIMIT ?",
            (biz_id, tur, chek, n)).fetchall()


def tez_kb(biz_id: int, tur: str) -> InlineKeyboardMarkup | None:
    rows = tez_yozuvlar(biz_id, tur)
    if not rows:
        return None
    tugma = []
    for i, r in enumerate(rows):
        belgi = "💵" if r["tolov"] == "Naqd" else ("💳" if r["tolov"] == "Karta" else "📝")
        pul = f"{fmt(r['summa'])} $" if r["valyuta"] == "USD" else fmt(r["summa"])
        tugma.append([InlineKeyboardButton(
            text=f"{belgi} {r['izoh'][:22]} · {pul}", callback_data=f"tez:{i}")])
    tugma.append([InlineKeyboardButton(text="✏️ Yangi summa yozish", callback_data="tez:yangi")])
    tugma.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=tugma)



@dp.message(F.text.in_({"➕ Kirim", "➖ Chiqim"}))
async def start_tx(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    kind = "Kirim" if "Kirim" in message.text else "Chiqim"
    bizlar = my_businesses(message.from_user.id)
    if not bizlar:
        await message.answer(_biznes_yoq(message.from_user.id))
        return
    await state.update_data(kind=kind)
    if len(bizlar) == 1:
        await state.update_data(biznes_id=bizlar[0]["id"])
        await state.set_state(TxForm.amount)
        kb = tez_kb(bizlar[0]["id"], kind)
        await message.answer(
            f"{biz_name(bizlar[0]['id'])} · <b>{kind}</b>\n\n"
            + ("Tez-tez yozadiganlaringiz yoki summani yozing:"
               if kb else "Summani yozing:\n"
                          "<code>150000</code> · <code>1,5 mln</code> · <code>20$</code>"),
            reply_markup=kb or cancel_kb())
        return
    await state.set_state(TxForm.business)
    await message.answer(f"<b>{kind}</b> — qaysi biznes?",
                         reply_markup=biz_kb(message.from_user.id, "tbiz"))


@dp.callback_query(TxForm.business, F.data.startswith("tbiz:"))
async def cb_tx_biz(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1])
    await state.update_data(biznes_id=biz_id)
    await state.set_state(TxForm.amount)
    data = await state.get_data()
    kb = tez_kb(biz_id, data["kind"])
    await cb.message.edit_text(
        f"{biz_name(biz_id)} · <b>{data['kind']}</b>\n\n"
        + ("Tez-tez yozadiganlaringiz yoki summani yozing:" if kb else
           "Summani yozing:\n<code>150000</code> · <code>1,5 mln</code> · <code>20$</code>"),
        reply_markup=kb)
    await cb.answer()


@dp.callback_query(F.data == "cancel")
async def cb_cancel(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("❌ Bekor qilindi.")
    await cb.answer()


@dp.callback_query(TxForm.amount, F.data.startswith("tez:"))
async def cb_tez(cb: CallbackQuery, state: FSMContext):
    tanlov = cb.data.split(":")[1]
    data = await state.get_data()
    if tanlov == "yangi":
        await cb.message.edit_text(
            f"{biz_name(data['biznes_id'])} · <b>{data['kind']}</b>\n\n"
            "Summani yozing:\n<code>150000</code> · <code>1,5 mln</code> · <code>20$</code>")
        await cb.answer()
        return
    rows = tez_yozuvlar(data["biznes_id"], data["kind"])
    try:
        r = rows[int(tanlov)]
    except (ValueError, IndexError):
        await cb.answer("Topilmadi", show_alert=True)
        return
    await state.clear()
    await cb.message.edit_text("💾 Saqlandi")
    await save_tx(cb.message, biznes_id=data["biznes_id"], kind=data["kind"],
                  amount=r["summa"], currency=r["valyuta"], note=r["izoh"],
                  payment=r["tolov"], uid=cb.from_user.id, author=uname(cb))
    await cb.answer()


@dp.message(TxForm.amount, F.text)
async def st_amount(message: Message, state: FSMContext):
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Summani tushunmadim. Masalan: <code>150000</code>")
        return
    await state.update_data(amount=parsed[0], currency=parsed[1])
    await state.set_state(TxForm.note)
    await message.answer("Izoh yozing (nima uchun / kim bilan):", reply_markup=note_kb())


@dp.callback_query(TxForm.note, F.data == "note_skip")
async def cb_skip_note(cb: CallbackQuery, state: FSMContext):
    await state.update_data(note="")
    await state.set_state(TxForm.payment)
    await cb.message.edit_text("To'lov turi:", reply_markup=tolov_kb())
    await cb.answer()


@dp.message(TxForm.note, F.text)
async def st_note(message: Message, state: FSMContext):
    await state.update_data(note=message.text.strip())
    await state.set_state(TxForm.payment)
    await message.answer("To'lov turi:", reply_markup=tolov_kb())


@dp.callback_query(TxForm.payment, F.data.startswith("pay:"))
async def cb_payment(cb: CallbackQuery, state: FSMContext):
    payment = TOLOVLAR[int(cb.data.split(":")[1])]
    data = await state.get_data()
    await state.clear()
    await cb.message.edit_text("💾 Saqlandi")
    await save_tx(cb.message, biznes_id=data["biznes_id"], kind=data["kind"],
                  amount=data["amount"], currency=data["currency"],
                  note=data.get("note", ""), payment=payment,
                  uid=cb.from_user.id, author=uname(cb))
    await cb.answer()


async def save_tx(target: Message, biznes_id: int, kind, amount, currency, note, payment,
                  uid: int | None = None, author: str | None = None,
                  sana: dt.date | None = None):
    uzs = amount * rate() if currency == "USD" else amount
    now = hozir()
    kun = sana or now.date()
    tx_id = tx_add(biznes_id=biznes_id, sana=kun.isoformat(),
                   vaqt=now.strftime("%H:%M"), user_id=uid or target.chat.id,
                   user_nomi=author or uname(target), tur=kind, izoh=note, tolov=payment,
                   valyuta=currency, summa=round(amount, 2), summa_uzs=round(uzs))
    inc, exp = totals(tx_period(kun.isoformat(), kun.isoformat(), biznes_id))
    bal = kassa_holati(biznes_id)
    icon = "🟢" if kind == "Kirim" else "🔴"
    money = f"{fmt(amount)} $" if currency == "USD" else f"{fmt(amount)} so'm"
    extra = f"  ({fmt(uzs)} so'm)" if currency == "USD" else ""
    tolov_str = {"Naqd": "💵 Naqd", "Karta": "💳 Karta",
                 "Qarzga": "📝 Qarzga (kassaga tegmadi)"}[payment]
    yozuvchi = uid or target.chat.id
    kassa_qismi = ""
    if is_admin(yozuvchi):
        ogoh = ""
        if payment in KASSALAR and bal[payment] < 0:
            ogoh = f"\n⚠️ <b>{payment} kassa minusda</b> — yozuv tushib qolganmi?"
        kassa_qismi = (f"\n💰 <b>Kassa:</b> 💵 {fmt(bal['Naqd'])} · "
                       f"💳 {fmt(bal['Karta'])} so'm{ogoh}\n")
    sana_qismi = "" if kun == bugun() else f"📅 <b>{kun:%d.%m.%Y}</b>\n"
    kun_nomi = "Bugun" if kun == bugun() else f"{kun:%d.%m}"
    await target.answer(
        f"{icon} <b>{kind} saqlandi</b>  <code>#{tx_id}</code>\n\n"
        f"🏢 {biz_name(biznes_id)}\n" + sana_qismi +
        f"💵 <b>{money}</b>{extra}\n"
        f"{tolov_str}\n" + (f"📝 {note}\n" if note else "") + kassa_qismi +
        f"<i>{kun_nomi}: kirim {fmt(inc)} · chiqim {fmt(exp)} · foyda {fmt(inc - exp)}</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="↩️ Bekor qilish", callback_data=f"undo:{tx_id}"),
            InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"edit:{tx_id}"),
        ]]))


# ── ↩️ bekor qilish ──────────────────────────────────────────────────
@dp.callback_query(F.data.startswith("undo:"))
async def cb_undo(cb: CallbackQuery):
    if not allowed(cb.from_user.id):
        return
    idlar = [int(x) for x in cb.data.split(":")[1].split(",") if x.strip().isdigit()]
    ochirildi = sum(1 for i in idlar if tx_delete(i, my_ids(cb.from_user.id)))
    if not ochirildi:
        await cb.answer("Yozuv topilmadi (allaqachon o'chirilgan)", show_alert=True)
        return
    await cb.message.edit_text(f"↩️ <b>Bekor qilindi</b> — {ochirildi} ta yozuv o'chirildi.")
    await cb.answer("Bekor qilindi")


# ── ✏️ tahrirlash ────────────────────────────────────────────────────
def tahrir_matni(r: sqlite3.Row) -> str:
    b = biz_get(r["biznes_id"])
    icon = "🟢" if r["tur"] == "Kirim" else "🔴"
    pul = f"{fmt(r['summa'])} $" if r["valyuta"] == "USD" else f"{fmt(r['summa_uzs'])} so'm"
    return (f"✏️ <b>Yozuvni tahrirlash</b>  <code>#{r['id']}</code>\n\n"
            f"{icon} {r['tur']} · <b>{pul}</b>\n"
            f"🏢 {b['emoji'] + ' ' + b['nomi'] if b else '—'}\n"
            f"📅 {r['sana']}  ⏰ {r['vaqt']}\n"
            f"💳 {r['tolov']}\n"
            f"📝 {r['izoh'] or '—'}\n\n"
            "Nimani o'zgartiramiz?")


def tahrir_kb(tx_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💵 Summa", callback_data=f"ed_sum:{tx_id}"),
         InlineKeyboardButton(text="📝 Izoh", callback_data=f"ed_izoh:{tx_id}")],
        [InlineKeyboardButton(text="💳 To'lov turi", callback_data=f"ed_tolov:{tx_id}"),
         InlineKeyboardButton(text="🏢 Biznes", callback_data=f"ed_biz:{tx_id}")],
        [InlineKeyboardButton(text="🔄 Kirim/Chiqim", callback_data=f"ed_tur:{tx_id}")],
        [InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"ed_och:{tx_id}")],
        [InlineKeyboardButton(text="✅ Yopish", callback_data="ed_yop")],
    ])


def tx_get(tx_id: int, biz_ids: list[int]) -> sqlite3.Row | None:
    if not biz_ids:
        return None
    marks = ",".join("?" * len(biz_ids))
    with db() as c:
        return c.execute(f"SELECT * FROM amaliyotlar WHERE id=? AND biznes_id IN ({marks})",
                         (tx_id, *biz_ids)).fetchone()


def tx_update(tx_id: int, **maydonlar) -> None:
    if not maydonlar:
        return
    qism = ", ".join(f"{k}=?" for k in maydonlar)
    with db() as c:
        c.execute(f"UPDATE amaliyotlar SET {qism} WHERE id=?",
                  (*maydonlar.values(), tx_id))


@dp.callback_query(F.data.startswith("edit:"))
async def cb_edit(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    r = tx_get(int(cb.data.split(":")[1]), my_ids(cb.from_user.id))
    if not r:
        await cb.answer("Yozuv topilmadi", show_alert=True)
        return
    await cb.message.answer(tahrir_matni(r), reply_markup=tahrir_kb(r["id"]))
    await cb.answer()


@dp.message(Command("tahrir"))
async def cmd_tahrir(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.clear()
    parts = message.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        oxirgi = tx_last(5, my_ids(message.from_user.id))
        if not oxirgi:
            await message.answer("Hali yozuv yo'q.")
            return
        rows = [[InlineKeyboardButton(
            text=f"#{r['id']} · {fmt(r['summa_uzs'])} · {(r['izoh'] or r['tur'])[:20]}",
            callback_data=f"edit:{r['id']}")] for r in oxirgi]
        await message.answer("Qaysi yozuvni tahrirlaymiz?\n<i>Yoki: /tahrir 78</i>",
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
        return
    r = tx_get(int(parts[1]), my_ids(message.from_user.id))
    if not r:
        await message.answer(f"#{parts[1]} topilmadi.")
        return
    await message.answer(tahrir_matni(r), reply_markup=tahrir_kb(r["id"]))


@dp.callback_query(F.data == "ed_yop")
async def cb_ed_yop(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("✅ Tahrirlash yopildi.")
    await cb.answer()


@dp.callback_query(F.data.startswith("ed_och:"))
async def cb_ed_och(cb: CallbackQuery):
    tx_id = int(cb.data.split(":")[1])
    ok = tx_delete(tx_id, my_ids(cb.from_user.id))
    await cb.message.edit_text(f"🗑 <code>#{tx_id}</code> o'chirildi." if ok else "Topilmadi.")
    await cb.answer()


@dp.callback_query(F.data.startswith("ed_tur:"))
async def cb_ed_tur(cb: CallbackQuery):
    tx_id = int(cb.data.split(":")[1])
    r = tx_get(tx_id, my_ids(cb.from_user.id))
    if not r:
        await cb.answer("Topilmadi", show_alert=True)
        return
    tx_update(tx_id, tur=("Chiqim" if r["tur"] == "Kirim" else "Kirim"))
    r = tx_get(tx_id, my_ids(cb.from_user.id))
    await cb.message.edit_text(tahrir_matni(r), reply_markup=tahrir_kb(tx_id))
    await cb.answer("O'zgartirildi")


@dp.callback_query(F.data.startswith("ed_tolov:"))
async def cb_ed_tolov(cb: CallbackQuery):
    tx_id = int(cb.data.split(":")[1])
    await cb.message.edit_text("To'lov turini tanlang:", reply_markup=InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(text="💵 Naqd", callback_data=f"edt:{tx_id}:Naqd"),
            InlineKeyboardButton(text="💳 Karta", callback_data=f"edt:{tx_id}:Karta")],
            [InlineKeyboardButton(text="📝 Qarzga", callback_data=f"edt:{tx_id}:Qarzga")]]))
    await cb.answer()


@dp.callback_query(F.data.startswith("edt:"))
async def cb_edt(cb: CallbackQuery):
    _, tx_id, tolov = cb.data.split(":")
    if not tx_get(int(tx_id), my_ids(cb.from_user.id)):
        await cb.answer("Topilmadi", show_alert=True)
        return
    tx_update(int(tx_id), tolov=tolov)
    r = tx_get(int(tx_id), my_ids(cb.from_user.id))
    await cb.message.edit_text(tahrir_matni(r), reply_markup=tahrir_kb(int(tx_id)))
    await cb.answer("O'zgartirildi")


@dp.callback_query(F.data.startswith("ed_biz:"))
async def cb_ed_biz(cb: CallbackQuery):
    tx_id = int(cb.data.split(":")[1])
    rows = [[InlineKeyboardButton(text=f"{b['emoji']} {b['nomi']}",
                                  callback_data=f"edb:{tx_id}:{b['id']}")]
            for b in my_businesses(cb.from_user.id)]
    await cb.message.edit_text("Qaysi biznesga o'tkazamiz?",
                               reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await cb.answer()


@dp.callback_query(F.data.startswith("edb:"))
async def cb_edb(cb: CallbackQuery):
    _, tx_id, biz = cb.data.split(":")
    if not tx_get(int(tx_id), my_ids(cb.from_user.id)):
        await cb.answer("Topilmadi", show_alert=True)
        return
    if int(biz) not in my_ids(cb.from_user.id):
        await cb.answer("Ruxsat yo'q", show_alert=True)
        return
    tx_update(int(tx_id), biznes_id=int(biz))
    r = tx_get(int(tx_id), my_ids(cb.from_user.id))
    await cb.message.edit_text(tahrir_matni(r), reply_markup=tahrir_kb(int(tx_id)))
    await cb.answer("Ko'chirildi")


@dp.callback_query(F.data.startswith("ed_sum:"))
async def cb_ed_sum(cb: CallbackQuery, state: FSMContext):
    tx_id = int(cb.data.split(":")[1])
    await state.set_state(EditForm.summa)
    await state.update_data(tx_id=tx_id)
    await cb.message.edit_text("Yangi summani yozing:\n"
                               "<code>450000</code> · <code>1,5 mln</code> · <code>20$</code>")
    await cb.answer()


@dp.message(EditForm.summa, F.text)
async def ed_sum_do(message: Message, state: FSMContext):
    data = await state.get_data()
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Summani tushunmadim. Masalan: <code>450000</code>")
        return
    await state.clear()
    r = tx_get(data["tx_id"], my_ids(message.from_user.id))
    if not r:
        await message.answer("Yozuv topilmadi.")
        return
    summa, valyuta = parsed
    uzs = summa * rate() if valyuta == "USD" else summa
    tx_update(r["id"], summa=round(summa, 2), summa_uzs=round(uzs), valyuta=valyuta)
    r = tx_get(r["id"], my_ids(message.from_user.id))
    await message.answer(tahrir_matni(r), reply_markup=tahrir_kb(r["id"]))


@dp.callback_query(F.data.startswith("ed_izoh:"))
async def cb_ed_izoh(cb: CallbackQuery, state: FSMContext):
    tx_id = int(cb.data.split(":")[1])
    await state.set_state(EditForm.izoh)
    await state.update_data(tx_id=tx_id)
    await cb.message.edit_text("Yangi izohni yozing:")
    await cb.answer()


@dp.message(EditForm.izoh, F.text)
async def ed_izoh_do(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    r = tx_get(data["tx_id"], my_ids(message.from_user.id))
    if not r:
        await message.answer("Yozuv topilmadi.")
        return
    tx_update(r["id"], izoh=message.text.strip()[:200])
    r = tx_get(r["id"], my_ids(message.from_user.id))
    await message.answer(tahrir_matni(r), reply_markup=tahrir_kb(r["id"]))


# ═══════════════════════════════════════════════════════ 9. KASSA
@dp.message(F.text == "💰 Kassa")
@dp.message(Command("kassa"))
async def kassa_entry(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    if not is_admin(message.from_user.id):
        await message.answer("Kassa bo'limi faqat bot egasida.",
                             reply_markup=main_menu(message.from_user.id))
        return
    await state.clear()
    await message.answer(kassa_report(my_ids(message.from_user.id)),
                         reply_markup=kassa_kb(is_admin(message.from_user.id)))


# ---- boshlang'ich qoldiq ----
@dp.callback_query(F.data == "kassa_init")
async def cb_kassa_init(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    bizlar = my_businesses(cb.from_user.id)
    if len(bizlar) == 1:
        await state.update_data(biznes_id=bizlar[0]["id"])
        await state.set_state(KassaInit.values)
        await cb.message.answer(_init_prompt(bizlar[0]["id"]), reply_markup=cancel_kb())
    else:
        await state.set_state(KassaInit.business)
        await cb.message.answer("Qaysi biznes kassasi?",
                                reply_markup=biz_kb(cb.from_user.id, "ibiz"))
    await cb.answer()


def _init_prompt(biz_id: int) -> str:
    return (f"{biz_name(biz_id)}\n\n"
            "Hozir <b>naqd</b> va <b>kartada</b> qancha pul bor? Vergul bilan yozing:\n\n"
            "<code>2000000, 1500000</code>\n\n"
            "<i>Shu raqamdan boshlab bot kassani o'zi yuritadi.</i>")


@dp.callback_query(KassaInit.business, F.data.startswith("ibiz:"))
async def cb_kassa_init_biz(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1])
    await state.update_data(biznes_id=biz_id)
    await state.set_state(KassaInit.values)
    await cb.message.edit_text(_init_prompt(biz_id))
    await cb.answer()


@dp.message(KassaInit.values, F.text)
async def kassa_init_do(message: Message, state: FSMContext):
    parts = [p.strip() for p in message.text.replace(";", ",").split(",")]
    vals = [parse_amount(p) if p and p not in ("0", "-") else (0.0, "UZS") for p in parts]
    if len(vals) < 2 or any(v is None for v in vals[:2]):
        await message.answer("Ikkita raqam yozing, vergul bilan: <code>2000000, 1500000</code>")
        return
    data = await state.get_data()
    await state.clear()
    naqd = vals[0][0] * (rate() if vals[0][1] == "USD" else 1)
    karta = vals[1][0] * (rate() if vals[1][1] == "USD" else 1)
    kassa_set_boshlangich(data["biznes_id"], round(naqd), round(karta))
    bal = kassa_holati(data["biznes_id"])
    await message.answer(
        f"✅ <b>{biz_name(data['biznes_id'])}</b> kassasi sozlandi.\n\n"
        f"💵 Naqd: <b>{fmt(bal['Naqd'])}</b>\n"
        f"💳 Karta: <b>{fmt(bal['Karta'])}</b>\n\n"
        "<i>Bu yerda mavjud yozuvlar ham hisobga olingan.</i>",
        reply_markup=main_menu(message.from_user.id))


# ---- kassani tekshirish ----
@dp.callback_query(F.data == "kassa_check")
async def cb_kassa_check(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    bizlar = my_businesses(cb.from_user.id)
    if len(bizlar) == 1:
        await state.update_data(biznes_id=bizlar[0]["id"])
        await state.set_state(KassaCheck.kassa)
        await cb.message.answer("Qaysi kassani tekshiramiz?", reply_markup=grid(
            [f"{KASSA_EMOJI[k]} {k}" for k in KASSALAR], "ck", 2))
    else:
        await state.set_state(KassaCheck.business)
        await cb.message.answer("Qaysi biznes kassasi?",
                                reply_markup=biz_kb(cb.from_user.id, "cbiz"))
    await cb.answer()


@dp.callback_query(KassaCheck.business, F.data.startswith("cbiz:"))
async def cb_check_biz(cb: CallbackQuery, state: FSMContext):
    await state.update_data(biznes_id=int(cb.data.split(":")[1]))
    await state.set_state(KassaCheck.kassa)
    await cb.message.edit_text("Qaysi kassani tekshiramiz?", reply_markup=grid(
        [f"{KASSA_EMOJI[k]} {k}" for k in KASSALAR], "ck", 2))
    await cb.answer()


@dp.callback_query(KassaCheck.kassa, F.data.startswith("ck:"))
async def cb_check_kassa(cb: CallbackQuery, state: FSMContext):
    kassa = KASSALAR[int(cb.data.split(":")[1])]
    data = await state.get_data()
    await state.update_data(kassa=kassa)
    await state.set_state(KassaCheck.amount)
    bal = kassa_holati(data["biznes_id"])[kassa]
    await cb.message.edit_text(
        f"{biz_name(data['biznes_id'])} · {KASSA_EMOJI[kassa]} <b>{kassa}</b>\n\n"
        f"Hisob bo'yicha: <b>{fmt(bal)}</b> so'm\n\n"
        "Haqiqiy summani sanab yozing:")
    await cb.answer()


@dp.message(KassaCheck.amount, F.text)
async def kassa_check_do(message: Message, state: FSMContext):
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Raqam yozing, masalan <code>2315000</code>")
        return
    data = await state.get_data()
    await state.clear()
    biz_id, kassa = data["biznes_id"], data["kassa"]
    haqiqiy = parsed[0] * (rate() if parsed[1] == "USD" else 1)
    hisob = kassa_holati(biz_id)[kassa]
    farq = haqiqiy - hisob

    cfg_set(f"kassa_tekshiruv_{biz_id}", bugun().isoformat())
    if abs(farq) < 1:
        await message.answer(
            f"✅ <b>Kassa to'g'ri.</b>\n{KASSA_EMOJI[kassa]} {kassa}: "
            f"<b>{fmt(hisob)}</b> so'm — farq yo'q.", reply_markup=main_menu(message.from_user.id))
        return

    now = hozir()
    tur = "Chiqim" if farq < 0 else "Kirim"
    izoh = "Kassa tekshiruvi — kamomad" if farq < 0 else "Kassa tekshiruvi — ortiqcha"
    tx_add(biznes_id=biz_id, sana=now.strftime("%Y-%m-%d"), vaqt=now.strftime("%H:%M"),
           user_id=message.from_user.id, user_nomi=uname(message), tur=tur, izoh=izoh,
           tolov=kassa, valyuta="UZS", summa=round(abs(farq)), summa_uzs=round(abs(farq)))
    yangi = kassa_holati(biz_id)[kassa]
    belgi = "−" if farq < 0 else "+"
    await message.answer(
        f"{'⚠️' if farq < 0 else 'ℹ️'} <b>Farq: {belgi}{fmt(abs(farq))} so'm</b> "
        f"({'kamomad' if farq < 0 else 'ortiqcha'})\n\n"
        f"«{izoh}» yozuvi qo'shildi.\n"
        f"💰 {KASSA_EMOJI[kassa]} {kassa} endi: <b>{fmt(yangi)}</b> so'm",
        reply_markup=main_menu(message.from_user.id))


# ---- pul ko'chirish ----
@dp.callback_query(F.data == "kassa_move")
async def cb_kassa_move(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    bizlar = my_businesses(cb.from_user.id)
    if len(bizlar) == 1:
        await state.update_data(biznes_id=bizlar[0]["id"])
        await state.set_state(KassaMove.yonalish)
        await cb.message.answer("Qayerdan qayerga?", reply_markup=move_kb(bizlar[0]["id"]))
    else:
        await state.set_state(KassaMove.business)
        await cb.message.answer("Qaysi biznesdan?", reply_markup=biz_kb(cb.from_user.id, "mbiz"))
    await cb.answer()


def move_kb(biz_id: int) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text="💵 Naqd → 💳 Karta", callback_data="mv:n2k")],
        [InlineKeyboardButton(text="💳 Karta → 💵 Naqd", callback_data="mv:k2n")],
    ]
    if len(biz_all()) > 1:
        rows.append([InlineKeyboardButton(text="🏢 Boshqa biznesga (naqd)",
                                          callback_data="mv:biz")])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.callback_query(KassaMove.business, F.data.startswith("mbiz:"))
async def cb_move_biz(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1])
    await state.update_data(biznes_id=biz_id)
    await state.set_state(KassaMove.yonalish)
    await cb.message.edit_text(f"{biz_name(biz_id)}\n\nQayerdan qayerga?",
                               reply_markup=move_kb(biz_id))
    await cb.answer()


@dp.callback_query(KassaMove.yonalish, F.data.startswith("mv:"))
async def cb_move_dir(cb: CallbackQuery, state: FSMContext):
    kod = cb.data.split(":")[1]
    data = await state.get_data()
    if kod == "biz":
        rows = [[InlineKeyboardButton(text=f"{b['emoji']} {b['nomi']}",
                                      callback_data=f"mto:{b['id']}")]
                for b in biz_all() if b["id"] != data["biznes_id"]]
        rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
        await state.set_state(KassaMove.target_biz)
        await cb.message.edit_text("Qaysi biznesga?",
                                   reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    else:
        qayerdan, qayerga = ("Naqd", "Karta") if kod == "n2k" else ("Karta", "Naqd")
        await state.update_data(qayerdan=qayerdan, qayerga=qayerga, biznes_2=None)
        await state.set_state(KassaMove.amount)
        await cb.message.edit_text(
            f"{KASSA_EMOJI[qayerdan]} {qayerdan} → {KASSA_EMOJI[qayerga]} {qayerga}\n\n"
            "Summani yozing:")
    await cb.answer()


@dp.callback_query(KassaMove.target_biz, F.data.startswith("mto:"))
async def cb_move_target(cb: CallbackQuery, state: FSMContext):
    await state.update_data(qayerdan="Naqd", qayerga="Naqd",
                            biznes_2=int(cb.data.split(":")[1]))
    await state.set_state(KassaMove.amount)
    data = await state.get_data()
    await cb.message.edit_text(
        f"{biz_name(data['biznes_id'])} → {biz_name(data['biznes_2'])} (naqd)\n\n"
        "Summani yozing:")
    await cb.answer()


@dp.message(KassaMove.amount, F.text)
async def kassa_move_do(message: Message, state: FSMContext):
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Raqam yozing, masalan <code>1000000</code>")
        return
    data = await state.get_data()
    await state.clear()
    summa = round(parsed[0] * (rate() if parsed[1] == "USD" else 1))
    oldin = kassa_holati(data["biznes_id"])
    now = hozir()
    kochirish_add(sana=now.strftime("%Y-%m-%d"), vaqt=now.strftime("%H:%M"),
                  user_nomi=uname(message), biznes_id=data["biznes_id"],
                  qayerdan=data["qayerdan"], biznes_2=data.get("biznes_2"),
                  qayerga=data["qayerga"], summa=summa)
    keyin = kassa_holati(data["biznes_id"])

    if data.get("biznes_2"):
        b2 = kassa_holati(data["biznes_2"])
        text = (f"✅ <b>Ko'chirildi</b>\n\n"
                f"{biz_name(data['biznes_id'])} → {biz_name(data['biznes_2'])}\n"
                f"💵 <b>{fmt(summa)}</b> so'm\n"
                f"─────────────\n"
                f"{biz_get(data['biznes_id'])['nomi']}: 💵 <b>{fmt(keyin['Naqd'])}</b>\n"
                f"{biz_get(data['biznes_2'])['nomi']}: 💵 <b>{fmt(b2['Naqd'])}</b>")
    else:
        qd, qg = data["qayerdan"], data["qayerga"]
        text = (f"✅ <b>Ko'chirildi</b>\n\n"
                f"{biz_name(data['biznes_id'])}\n"
                f"{KASSA_EMOJI[qd]} {qd} → {KASSA_EMOJI[qg]} {qg}: <b>{fmt(summa)}</b> so'm\n"
                f"─────────────\n"
                f"💵 Naqd:  {fmt(oldin['Naqd'])} → <b>{fmt(keyin['Naqd'])}</b>\n"
                f"💳 Karta: {fmt(oldin['Karta'])} → <b>{fmt(keyin['Karta'])}</b>")
    await message.answer(text + "\n\n<i>Bu kirim ham, chiqim ham emas — foydaga ta'sir qilmaydi.</i>",
                         reply_markup=main_menu(message.from_user.id))


# ---- kassa harakati ----
@dp.callback_query(F.data == "kassa_hist")
async def cb_kassa_hist(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    ids = my_ids(cb.from_user.id)
    if not ids:
        await cb.answer("Biznes yo'q", show_alert=True)
        return
    for i in ids[:3]:
        await cb.message.answer(kassa_harakati(i))
    await cb.answer()


# ═══════════════════════════════════════════════════════ 10. HISOBOTLAR
MONTHS_UZ = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
             "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]


@dp.message(F.text == "📊 Bugun")
async def rep_today(message: Message):
    if not allowed(message.from_user.id):
        return
    t = today_iso()
    await message.answer(report(t, t, f"📊 BUGUN — {bugun():%d.%m.%Y}",
                                my_ids(message.from_user.id),
                                is_admin(message.from_user.id)),
                         reply_markup=main_menu(message.from_user.id))


@dp.message(Command("hafta"))
async def rep_week(message: Message):
    if not allowed(message.from_user.id):
        return
    end = bugun()
    start = end - dt.timedelta(days=6)
    await message.answer(report(start.isoformat(), end.isoformat(),
                                f"📆 OXIRGI 7 KUN ({start:%d.%m} – {end:%d.%m})",
                                my_ids(message.from_user.id),
                                is_admin(message.from_user.id)),
                         reply_markup=main_menu(message.from_user.id))


@dp.message(F.text == "📅 Shu oy")
async def rep_month(message: Message):
    if not allowed(message.from_user.id):
        return
    t = bugun()
    ids = my_ids(message.from_user.id)
    text = report(t.replace(day=1).isoformat(), t.isoformat(),
                  f"📅 {MONTHS_UZ[t.month - 1].upper()} {t.year}", ids,
                  is_admin(message.from_user.id))
    rows = [r for r in tx_period(t.replace(day=1).isoformat(), t.isoformat())
            if r["biznes_id"] in ids]
    by_day: dict[str, float] = defaultdict(float)
    for r in rows:
        by_day[r["sana"]] += r["summa_uzs"] * (1 if r["tur"] == "Kirim" else -1)
    if by_day:
        best = max(by_day.items(), key=lambda kv: kv[1])
        text += (f"\n\n🏆 Eng foydali kun: {best[0]} — {fmt(best[1])} so'm"
                 f"\n📈 O'rtacha kunlik foyda: {fmt(sum(by_day.values()) / len(by_day))} so'm")
    await message.answer(text, reply_markup=main_menu(message.from_user.id))


@dp.message(Command("yil"))
async def rep_year(message: Message):
    if not allowed(message.from_user.id):
        return
    t = bugun()
    ids = my_ids(message.from_user.id)
    text = report(f"{t.year}-01-01", t.isoformat(), f"🗓 {t.year}-YIL", ids,
                  is_admin(message.from_user.id))
    lines = ["", "<b>Oylar kesimida (umumiy foyda):</b>"]
    for m in range(1, t.month + 1):
        start = dt.date(t.year, m, 1)
        end = dt.date(t.year + (m == 12), (m % 12) + 1, 1) - dt.timedelta(days=1)
        rows = [r for r in tx_period(start.isoformat(), end.isoformat()) if r["biznes_id"] in ids]
        inc, exp = totals(rows)
        lines.append(f"  {MONTHS_UZ[m - 1][:3]}: {fmt(inc - exp)} so'm")
    await message.answer(text + "\n" + "\n".join(lines), reply_markup=main_menu(message.from_user.id))


@dp.message(Command("ochirish"))
async def cmd_delete(message: Message):
    if not allowed(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: <code>/ochirish 12</code>")
        return
    ok = tx_delete(int(parts[1]), my_ids(message.from_user.id))
    await message.answer(f"🗑 #{parts[1]} o'chirildi." if ok else f"#{parts[1]} topilmadi.")


# ── eski xato yozuvlarni topish (150.000 -> 150 muammosi) ──────────────
def shubhali_yozuvlar(biz_ids: list[int]) -> list[sqlite3.Row]:
    """UZS da 1000 so'mdan kichik yozuvlar — nuqta xatosi qurboni bo'lishi mumkin."""
    if not biz_ids:
        return []
    marks = ",".join("?" * len(biz_ids))
    with db() as c:
        return c.execute(
            f"SELECT * FROM amaliyotlar WHERE biznes_id IN ({marks}) "
            "AND valyuta='UZS' AND summa_uzs > 0 AND summa_uzs < 1000 ORDER BY id",
            tuple(biz_ids)).fetchall()


@dp.message(Command("shubhali"))
async def cmd_shubhali(message: Message):
    if not is_admin(message.from_user.id):
        return
    rows = shubhali_yozuvlar(my_ids(message.from_user.id))
    if not rows:
        await message.answer("✅ Shubhali yozuv topilmadi — hammasi joyida.",
                             reply_markup=main_menu(message.from_user.id))
        return
    lines = ["🔍 <b>SHUBHALI YOZUVLAR</b>", "",
             "Bu yozuvlar 1000 so'mdan kichik. Ehtimol <code>150.000</code> ko'rinishida "
             "yozilgan va eski xato tufayli 1000 barobar kichrayib qolgan.", ""]
    for r in rows[:25]:
        b = biz_get(r["biznes_id"])
        icon = "🟢" if r["tur"] == "Kirim" else "🔴"
        lines.append(f"{icon} <code>#{r['id']}</code> {b['emoji'] if b else ''} {r['sana']} · "
                     f"<b>{fmt(r['summa_uzs'])}</b> so'm"
                     + (f" · {r['izoh']}" if r["izoh"] else ""))
    if len(rows) > 25:
        lines.append(f"<i>... va yana {len(rows) - 25} ta</i>")
    lines.append("")
    lines.append("Hammasini 1000 ga ko'paytirib tuzatamizmi?")
    await message.answer("\n".join(lines), reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"🔧 Hammasini ×1000 ({len(rows)} ta)",
                              callback_data="fix1000")],
        [InlineKeyboardButton(text="✋ Tegmang", callback_data="fix_yoq")],
    ]))


@dp.callback_query(F.data == "fix_yoq")
async def cb_fix_yoq(cb: CallbackQuery):
    await cb.message.edit_text("✋ Hech narsa o'zgartirilmadi.")
    await cb.answer()


@dp.callback_query(F.data == "fix1000")
async def cb_fix1000(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    rows = shubhali_yozuvlar(my_ids(cb.from_user.id))
    if not rows:
        await cb.message.edit_text("Shubhali yozuv qolmadi.")
        await cb.answer()
        return
    with db() as c:
        for r in rows:
            c.execute("UPDATE amaliyotlar SET summa=summa*1000, summa_uzs=summa_uzs*1000, "
                      "izoh=TRIM(COALESCE(izoh,'') || ' (tuzatildi ×1000)') WHERE id=?",
                      (r["id"],))
    await cb.message.edit_text(
        f"✅ <b>{len(rows)} ta yozuv tuzatildi</b> (×1000).\n\n"
        "Endi 💰 Kassa → 🧮 Kassani tekshirish orqali qoldiqni haqiqiy pulga tenglashtiring.")
    await cb.answer()


# ═══════════════════════════════════════════════════════ 11. QARZLAR
@dp.message(F.text == "💳 Qarzlar")
async def debts_entry(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.clear()
    await message.answer("💳 <b>Qarzlar bo'limi</b>", reply_markup=debts_kb())


@dp.callback_query(F.data == "debt_list")
async def cb_debt_list(cb: CallbackQuery):
    await cb.message.edit_text(debts_report(my_ids(cb.from_user.id)), reply_markup=debts_kb())
    await cb.answer()


@dp.callback_query(F.data.startswith("debt_new:"))
async def cb_debt_new(cb: CallbackQuery, state: FSMContext):
    kind = "Men qarz berdim" if cb.data.endswith("out") else "Men qarz oldim"
    await state.update_data(turi=kind)
    bizlar = my_businesses(cb.from_user.id)
    if len(bizlar) == 1:
        await state.update_data(biznes_id=bizlar[0]["id"])
        await state.set_state(DebtForm.who)
        await cb.message.edit_text(f"<b>{kind}</b>\n\nKimga / kimdan? Ism yozing:",
                                   reply_markup=cancel_kb())
    else:
        await state.set_state(DebtForm.business)
        await cb.message.edit_text(f"<b>{kind}</b>\n\nQaysi biznes bo'yicha?",
                                   reply_markup=biz_kb(cb.from_user.id, "dbiz"))
    await cb.answer()


@dp.callback_query(DebtForm.business, F.data.startswith("dbiz:"))
async def cb_debt_biz(cb: CallbackQuery, state: FSMContext):
    await state.update_data(biznes_id=int(cb.data.split(":")[1]))
    await state.set_state(DebtForm.who)
    data = await state.get_data()
    await cb.message.edit_text(
        f"{biz_name(data['biznes_id'])} · <b>{data['turi']}</b>\n\nKimga / kimdan? Ism yozing:",
        reply_markup=cancel_kb())
    await cb.answer()


@dp.message(DebtForm.who, F.text)
async def debt_who(message: Message, state: FSMContext):
    await state.update_data(kim=message.text.strip())
    await state.set_state(DebtForm.amount)
    await message.answer("Summani yozing (<code>2000000</code> yoki <code>200$</code>):")


@dp.message(DebtForm.amount, F.text)
async def debt_amount(message: Message, state: FSMContext):
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Summani tushunmadim.")
        return
    await state.update_data(amount=parsed[0], currency=parsed[1])
    await state.set_state(DebtForm.note)
    await message.answer(
        "Izoh yoki muddat (kerak bo'lmasa <code>-</code>):\n\n"
        "<i>Muddat yozsangiz o'sha kuni eslataman:</i> <code>01.09 mart oyigacha</code>")


@dp.message(DebtForm.note, F.text)
async def debt_note(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    note = "" if message.text.strip() == "-" else message.text.strip()
    muddat, qolgan = sanani_top(note) if note else (None, "")
    if muddat is None and note:            # kelajakdagi sana ham bo'lishi mumkin
        m2 = SANA_RE.match(note.split()[0]) if note.split() else None
        if m2:
            try:
                yil = int(m2.group(3)) + (2000 if m2.group(3) and int(m2.group(3)) < 100 else 0) \
                    if m2.group(3) else bugun().year
                muddat = dt.date(yil, int(m2.group(2)), int(m2.group(1)))
                qolgan = note[len(note.split()[0]):].strip()
            except ValueError:
                muddat = None
    if muddat:
        note = qolgan
    amount, currency = data["amount"], data["currency"]
    uzs = amount * rate() if currency == "USD" else amount
    debt_id = debt_add(biznes_id=data["biznes_id"], sana=today_iso(), user_nomi=uname(message),
                       kim=data["kim"], turi=data["turi"], valyuta=currency,
                       summa=round(amount, 2), summa_uzs=round(uzs), izoh=note,
                       muddat=muddat.isoformat() if muddat else None)
    muddat_str = f"\n📅 Muddat: <b>{muddat:%d.%m.%Y}</b> — o'sha kuni eslataman" if muddat else ""
    await message.answer(
        f"💳 <b>Qarz yozildi</b>  <code>#{debt_id}</code>\n\n"
        f"🏢 {biz_name(data['biznes_id'])}\n"
        f"{data['turi']}\n👤 {data['kim']}: <b>{fmt(uzs)} so'm</b>{muddat_str}\n\n"
        "<i>Kassaga hozir tegmadi. Qarz yopilganda pul kassaga kiritiladi.</i>",
        reply_markup=main_menu(message.from_user.id))


@dp.callback_query(F.data == "debt_close")
async def cb_debt_close(cb: CallbackQuery, state: FSMContext):
    await state.set_state(CloseDebt.debt_id)
    await cb.message.edit_text(
        "Yopiladigan qarz raqamini yozing (masalan <code>3</code>).\n"
        "Raqamlarni «📋 Ro'yxat» dan ko'rasiz.", reply_markup=cancel_kb())
    await cb.answer()


@dp.message(CloseDebt.debt_id, F.text)
async def debt_close_ask(message: Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("Faqat raqam yuboring.")
        return
    d = debt_get(int(message.text.strip()), my_ids(message.from_user.id))
    if not d:
        await state.clear()
        await message.answer("Bunday raqamli qarz topilmadi.", reply_markup=main_menu(message.from_user.id))
        return
    await state.update_data(debt_id=d["id"])
    await state.set_state(CloseDebt.kassaga)
    qaysi = "kirim" if d["turi"] == "Men qarz berdim" else "chiqim"
    await message.answer(
        f"<code>#{d['id']}</code> {d['kim']} — <b>{fmt(d['summa_uzs'])}</b> so'm\n"
        f"{d['turi']}\n\nPul qaysi kassa orqali o'tdi?\n"
        f"<i>Tanlasangiz, kassaga {qaysi} bo'lib tushadi.</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💵 Naqd", callback_data="dk:Naqd"),
             InlineKeyboardButton(text="💳 Karta", callback_data="dk:Karta")],
            [InlineKeyboardButton(text="🚫 Kassaga yozilmasin", callback_data="dk:yoq")],
        ]))


@dp.callback_query(CloseDebt.kassaga, F.data.startswith("dk:"))
async def cb_debt_close_do(cb: CallbackQuery, state: FSMContext):
    tanlov = cb.data.split(":")[1]
    data = await state.get_data()
    await state.clear()
    d = debt_get(data["debt_id"], my_ids(cb.from_user.id))
    if not d:
        await cb.message.edit_text("Qarz topilmadi.")
        await cb.answer()
        return
    debt_close(d["id"])
    qoshimcha = ""
    if tanlov in KASSALAR:
        now = hozir()
        tur = "Kirim" if d["turi"] == "Men qarz berdim" else "Chiqim"
        tx_add(biznes_id=d["biznes_id"], sana=now.strftime("%Y-%m-%d"),
               vaqt=now.strftime("%H:%M"), user_id=cb.from_user.id, user_nomi=uname(cb),
               tur=tur, izoh=f"Qarz yopildi — {d['kim']}", tolov=tanlov,
               valyuta="UZS", summa=round(d["summa_uzs"]), summa_uzs=round(d["summa_uzs"]))
        bal = kassa_holati(d["biznes_id"])
        qoshimcha = (f"\n\n{KASSA_EMOJI[tanlov]} Kassaga {tur.lower()} qilib yozildi.\n"
                     f"💰 💵 {fmt(bal['Naqd'])} · 💳 {fmt(bal['Karta'])} so'm")
    await cb.message.edit_text(f"✅ <code>#{d['id']}</code> qarz yopildi.{qoshimcha}")
    await cb.answer()


# ═══════════════════════════════════════════════════════ 12. EXCEL
@dp.message(F.text == "📥 Excel")
@dp.message(Command("excel"))
async def send_excel(message: Message):
    if not allowed(message.from_user.id):
        return
    ids = my_ids(message.from_user.id)
    jami = sum(len(tx_all(i)) for i in ids)
    if not jami:
        await message.answer("Hali birorta yozuv yo'q.", reply_markup=main_menu(message.from_user.id))
        return
    await message.answer("📥 Fayl tayyorlanmoqda...")
    data = await asyncio.to_thread(build_excel, ids, is_admin(message.from_user.id))
    await message.answer_document(
        BufferedInputFile(data, filename=f"Hisob-kitob-{bugun():%Y-%m-%d}.xlsx"),
        caption=f"📊 {len(ids)} ta biznes, {jami} ta amaliyot.\n"
                "Har biznesga alohida varaq + Xulosa (kassa qoldig'i bilan).",
        reply_markup=main_menu(message.from_user.id))


# ═══════════════════════════════════════════════════════ 13. SOZLAMALAR
@dp.message(F.text == "⚙️ Sozlamalar")
async def settings(message: Message, state: FSMContext):
    if not allowed(message.from_user.id):
        return
    await state.clear()
    u = user_get(message.from_user.id)
    ro = "\n".join(f"  {b['emoji']} {b['nomi']} — {len(tx_all(b['id']))} ta yozuv"
                   for b in my_businesses(message.from_user.id)) or "  —"
    await message.answer(
        f"⚙️ <b>Sozlamalar</b>\n\n"
        f"👤 Siz: {u['nomi']} ({u['rol']})\n"
        f"🆔 ID: <code>{message.from_user.id}</code>\n"
        f"💱 USD kursi: <b>{fmt(rate())}</b> so'm\n\n"
        f"<b>Bizneslar:</b>\n{ro}",
        reply_markup=settings_kb(is_admin(message.from_user.id)))


@dp.callback_query(F.data == "last10")
async def cb_last10(cb: CallbackQuery):
    rows = tx_last(10, my_ids(cb.from_user.id))
    if not rows:
        await cb.answer("Hali yozuv yo'q", show_alert=True)
        return
    lines = ["🧾 <b>Oxirgi 10 ta amaliyot</b>", ""]
    for r in rows:
        icon = "🟢" if r["tur"] == "Kirim" else "🔴"
        b = biz_get(r["biznes_id"])
        izoh = f" · {r['izoh']}" if r["izoh"] else ""
        lines.append(f"{icon} <code>#{r['id']}</code> {b['emoji'] if b else ''} {r['sana']} · "
                     f"<b>{fmt(r['summa_uzs'])}</b> · {r['tolov']}{izoh}")
    lines.append("\n<i>O'chirish: /ochirish RAQAM</i>")
    await cb.message.answer("\n".join(lines))
    await cb.answer()


@dp.callback_query(F.data == "biz_menu")
async def cb_biz_menu(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    lines = ["🏢 <b>Bizneslar</b>", ""]
    for b in biz_all():
        rows = tx_all(b["id"])
        inc, exp = totals(rows)
        bal = kassa_holati(b["id"])
        lines.append(f"{b['emoji']} <b>{b['nomi']}</b>")
        lines.append(f"   {len(rows)} ta yozuv · foyda {fmt(inc - exp)} so'm")
        lines.append(f"   💵 {fmt(bal['Naqd'])} · 💳 {fmt(bal['Karta'])}")
    lines.append("\n<i>Nomini o'zgartirish uchun ustiga bosing.</i>")
    await cb.message.answer("\n".join(lines), reply_markup=biz_menu_kb())
    await cb.answer()


@dp.callback_query(F.data == "biz_new")
async def cb_biz_new(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    await state.set_state(BizForm.new_name)
    await cb.message.answer("Yangi biznes nomini yozing:", reply_markup=cancel_kb())
    await cb.answer()


@dp.message(BizForm.new_name, F.text)
async def biz_new_do(message: Message, state: FSMContext):
    nomi = message.text.strip()
    if not nomi or nomi in MENYU_TUGMALARI:
        await message.answer("Biznes nomini yozing (menyu tugmasi emas):")
        return
    await state.clear()
    biz_add(nomi)
    await message.answer(f"✅ <b>{nomi}</b> qo'shildi.\n\n"
                         "💰 Kassa → ⚙️ Boshlang'ich qoldiq dan kassasini sozlang.",
                         reply_markup=main_menu(message.from_user.id))


@dp.callback_query(F.data.startswith("biz_ren:"))
async def cb_biz_rename(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    biz_id = int(cb.data.split(":")[1])
    await state.set_state(BizForm.rename)
    await state.update_data(biznes_id=biz_id)
    await cb.message.answer(
        f"{biz_name(biz_id)} — yangi nomni yozing.\n\n"
        "<i>Emoji ham o'zgartirmoqchi bo'lsangiz, nom oldiga yozing: "
        "<code>🚗 Avtoservis</code></i>", reply_markup=cancel_kb())
    await cb.answer()


@dp.message(BizForm.rename, F.text)
async def biz_rename_do(message: Message, state: FSMContext):
    text = message.text.strip()
    if not text or text in MENYU_TUGMALARI:
        await message.answer("Yangi nomni yozing (menyu tugmasi emas):")
        return
    data = await state.get_data()
    await state.clear()
    parts = text.split(maxsplit=1)
    if len(parts) == 2 and not parts[0][:1].isalnum() and len(parts[0]) <= 3:
        biz_set_emoji(data["biznes_id"], parts[0])
        text = parts[1].strip()
    biz_rename(data["biznes_id"], text)
    await message.answer(f"✅ O'zgartirildi: <b>{biz_name(data['biznes_id'])}</b>",
                         reply_markup=main_menu(message.from_user.id))


@dp.callback_query(F.data == "zaxira_ol")
async def cb_zaxira(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    await cb.answer("Tayyorlanmoqda...")
    await zaxira_yubor(cb.from_user.id, "qo'lda so'ralgan")


@dp.callback_query(F.data == "kalitlar")
async def cb_kalitlar(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    lines = ["🔑 <b>KALIT SO'ZLAR</b>", "",
             "Izohda shu so'zlar uchrasa, bot biznesni <b>o'zi tanlaydi</b> — "
             "so'ramaydi.", ""]
    rows = []
    for b in biz_all():
        sozlar = [r["soz"] for r in kalit_all(b["id"])]
        lines.append(f"{b['emoji']} <b>{b['nomi']}</b>")
        lines.append("   " + (", ".join(sozlar) if sozlar else "<i>hali yo'q</i>"))
        rows.append([InlineKeyboardButton(text=f"✏️ {b['emoji']} {b['nomi']}",
                                          callback_data=f"kalit_ed:{b['id']}")])
    lines.append("")
    lines.append("<i>Masalan: savdo, tovar, kassa</i>")
    await cb.message.answer("\n".join(lines),
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    await cb.answer()


@dp.callback_query(F.data.startswith("kalit_ed:"))
async def cb_kalit_ed(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    biz_id = int(cb.data.split(":")[1])
    await state.set_state(KalitForm.sozlar)
    await state.update_data(biznes_id=biz_id)
    bor = ", ".join(r["soz"] for r in kalit_all(biz_id)) or "—"
    await cb.message.answer(
        f"{biz_name(biz_id)} uchun kalit so'zlar.\n\n"
        f"Hozirgilari: <code>{bor}</code>\n\n"
        "Yangi ro'yxatni vergul bilan yozing (eskisi almashadi):\n"
        "<code>savdo, tovar, kassa, sotildi</code>\n\n"
        "<i>Hech qanday so'z kerak bo'lmasa:</i> <code>-</code>",
        reply_markup=cancel_kb())
    await cb.answer()


@dp.message(KalitForm.sozlar, F.text)
async def kalit_saqla(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    biz_id = data["biznes_id"]
    kalit_clear(biz_id)
    matn = message.text.strip()
    if matn == "-":
        await message.answer(f"✅ {biz_name(biz_id)} uchun kalit so'zlar o'chirildi.",
                             reply_markup=main_menu(message.from_user.id))
        return
    n = kalit_add(biz_id, matn)
    sozlar = ", ".join(r["soz"] for r in kalit_all(biz_id))
    await message.answer(
        f"✅ {biz_name(biz_id)} — <b>{n} ta</b> kalit so'z saqlandi:\n<code>{sozlar}</code>\n\n"
        "<i>Endi izohda shu so'zlar uchrasa, bot biznesni o'zi tanlaydi.</i>",
        reply_markup=main_menu(message.from_user.id))


@dp.callback_query(F.data == "set_rate")
async def cb_set_rate(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi o'zgartira oladi", show_alert=True)
        return
    await state.set_state(RateForm.value)
    await cb.message.answer("Yangi USD kursini yozing (masalan <code>12900</code>):")
    await cb.answer()


@dp.message(RateForm.value, F.text)
async def set_rate_do(message: Message, state: FSMContext):
    await state.clear()
    parsed = parse_amount(message.text)
    if parsed is None:
        await message.answer("Raqam yuboring, masalan <code>12900</code>")
        return
    cfg_set("usd_kurs", parsed[0])
    await message.answer(f"✅ USD kursi <b>{fmt(parsed[0])}</b> so'm qilib o'zgartirildi.",
                         reply_markup=main_menu(message.from_user.id))


@dp.callback_query(F.data == "users")
async def cb_users(cb: CallbackQuery):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    lines = ["👥 <b>Foydalanuvchilar</b>", ""]
    for u in users_all():
        if u["rol"] == "admin":
            lines.append(f"👑 {u['nomi']} — <code>{u['user_id']}</code> · hamma biznes")
        else:
            b = biz_get(u["biznes_id"]) if u["biznes_id"] else None
            joy = f"{b['emoji']} {b['nomi']}" if b else "hamma biznes"
            lines.append(f"👤 {u['nomi']} — <code>{u['user_id']}</code> · {joy}")
    lines.append("\n<i>O'chirish: /ochir_xodim ID</i>")
    await cb.message.answer("\n".join(lines), reply_markup=InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="➕ Xodim qo'shish",
                                               callback_data="add_user")]]))
    await cb.answer()


@dp.callback_query(F.data == "add_user")
async def cb_add_user(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("Faqat bot egasi", show_alert=True)
        return
    await state.set_state(AddUser.user_id)
    await cb.message.answer(
        "Xodimning <b>Telegram ID</b> raqamini yuboring.\n\n"
        "<i>Xodim botga /start yozsa, bot unga ID sini aytadi.</i>",
        reply_markup=cancel_kb())
    await cb.answer()


@dp.message(AddUser.user_id, F.text)
async def add_user_id(message: Message, state: FSMContext):
    text = message.text.strip()
    if not text.lstrip("-").isdigit():
        await message.answer("Faqat raqam yuboring.")
        return
    await state.update_data(new_uid=int(text))
    await state.set_state(AddUser.business)
    rows = [[InlineKeyboardButton(text=f"{b['emoji']} {b['nomi']}",
                                  callback_data=f"ubiz:{b['id']}")] for b in biz_all()]
    rows.append([InlineKeyboardButton(text="🌐 Hamma biznes", callback_data="ubiz:0")])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")])
    await message.answer("Bu xodim qaysi biznesda ishlaydi?",
                         reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@dp.callback_query(AddUser.business, F.data.startswith("ubiz:"))
async def cb_add_user_biz(cb: CallbackQuery, state: FSMContext):
    biz_id = int(cb.data.split(":")[1]) or None
    data = await state.get_data()
    await state.clear()
    uid = data["new_uid"]
    user_add(uid, f"Xodim {uid}", "xodim", biz_id)
    joy = biz_name(biz_id) if biz_id else "hamma biznes"
    await cb.message.edit_text(f"✅ <code>{uid}</code> qo'shildi.\n🏢 {joy}\n\n"
                               "U botga /start yozsa, ishlay boshlaydi.")
    await cb.answer()


@dp.message(Command("ochir_xodim"))
async def cmd_remove_user(message: Message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) < 2 or not parts[1].lstrip("-").isdigit():
        await message.answer("Foydalanish: <code>/ochir_xodim 123456789</code>")
        return
    ok = user_remove(int(parts[1]))
    await message.answer(f"🗑 <code>{parts[1]}</code> o'chirildi." if ok
                         else "Topilmadi (yoki u admin).")


# ═══════════════════════════════════════════════════════ 14. QOLGANI
# ── matn o'rniga stiker/rasm kelsa ──────────────────────────────────────
_MATN_KUTILADI = (
    Setup.names, TxForm.amount, TxForm.note, KassaInit.values, KassaCheck.amount,
    KassaMove.amount, DebtForm.who, DebtForm.amount, DebtForm.note, CloseDebt.debt_id,
    RateForm.value, AddUser.user_id, BizForm.new_name, BizForm.rename,
    EditForm.summa, EditForm.izoh, KalitForm.sozlar,
)


# ═══════════════════════════════════════════════════════ 15b-2. TIKLASH
_tikla_kutilmoqda: set[int] = set()


@dp.message(Command("tikla"))
async def cmd_tikla(message: Message):
    """Zaxira faylni yuborib, bazani to'liq tiklash."""
    if not is_admin(message.from_user.id):
        return
    _tikla_kutilmoqda.add(message.from_user.id)
    await message.answer(
        "♻️ <b>Bazani tiklash</b>\n\n"
        "Endi zaxira faylni (<code>.db</code>) shu yerga <b>fayl sifatida</b> "
        "yuboring.\n\n"
        "⚠️ Serverdagi hozirgi ma'lumotlar shu fayl bilan <b>almashtiriladi</b>. "
        "Ehtiyot bo'ling — avval /zaxira qilib, hozirgi holatni saqlab qo'ying.\n\n"
        "Bekor qilish: /bekor")


@dp.message(Command("bekor"))
async def cmd_bekor_tikla(message: Message):
    if message.from_user.id in _tikla_kutilmoqda:
        _tikla_kutilmoqda.discard(message.from_user.id)
        await message.answer("Bekor qilindi.")


@dp.message(F.document)
async def hujjat_qabul(message: Message):
    """Tiklash rejimida yuborilgan .db faylni qabul qiladi."""
    uid = message.from_user.id
    if uid not in _tikla_kutilmoqda or not is_admin(uid):
        return
    doc = message.document
    if not (doc.file_name or "").lower().endswith(".db"):
        await message.answer("Bu .db fayl emas. Zaxira faylni yuboring yoki /bekor.")
        return
    if doc.file_size and doc.file_size > 40 * 1024 * 1024:
        await message.answer("Fayl juda katta (40 MB dan ortiq).")
        return

    await message.answer("⏳ Tekshirilmoqda...")
    import tempfile
    tmpdir = tempfile.mkdtemp()
    yangi = Path(tmpdir) / "yangi.db"
    try:
        fayl = await bot.get_file(doc.file_id)
        await bot.download_file(fayl.file_path, destination=str(yangi))

        # Haqiqiy hisobot bazasimi?
        tekshir = sqlite3.connect(yangi)
        nomlar = {r[0] for r in tekshir.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        kerak = {"amaliyotlar", "bizneslar", "foydalanuvchilar", "sozlamalar"}
        if not kerak.issubset(nomlar):
            tekshir.close()
            await message.answer(
                "❌ Bu fayl hisob-kitob bazasiga o'xshamaydi.\n"
                f"Kerakli jadvallar topilmadi: {', '.join(sorted(kerak - nomlar))}")
            return
        n_tx = tekshir.execute("SELECT COUNT(*) FROM amaliyotlar").fetchone()[0]
        n_biz = tekshir.execute("SELECT COUNT(*) FROM bizneslar").fetchone()[0]
        tekshir.close()

        # Hozirgi holatni saqlab qo'yamiz (qaytarib bo'lishi uchun)
        eski = DB_PATH.with_name(f"hisobot-eski-{hozir():%Y%m%d-%H%M%S}.db")
        try:
            with db() as manba, sqlite3.connect(eski) as maqsad:
                manba.backup(maqsad)
        except Exception as e:
            log.warning("Eski nusxa saqlanmadi: %s", e)

        # Almashtiramiz
        global _wal_qilindi
        for qoshimcha in ("-wal", "-shm"):
            yon = Path(str(DB_PATH) + qoshimcha)
            if yon.exists():
                try:
                    yon.unlink()
                except OSError:
                    pass
        shutil.copy2(yangi, DB_PATH)
        _wal_qilindi = False
        init_db()

        _tikla_kutilmoqda.discard(uid)
        await message.answer(
            f"✅ <b>Baza tiklandi.</b>\n\n"
            f"{n_tx} ta amaliyot · {n_biz} ta biznes\n\n"
            f"Eski holat saqlandi: <code>{eski.name}</code>",
            reply_markup=main_menu(uid))
    except Exception as e:
        log.exception("Tiklashda xato")
        await message.answer(f"❌ Tiklab bo'lmadi: {e}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@dp.message(StateFilter(*_MATN_KUTILADI))
async def matn_kerak(message: Message):
    await message.answer("Iltimos, <b>matn</b> yozing (rasm yoki stiker emas).\n"
                         "Bekor qilish uchun /start yuboring.")


@dp.message()
async def fallback(message: Message):
    if not allowed(message.from_user.id):
        await message.answer(
            f"⛔️ Sizda ruxsat yo'q.\nSizning ID: <code>{message.from_user.id}</code>")
        return
    await message.answer(
        "Tushunmadim 🤔\n\n"
        "Tez yozish: <code>+150000 savdo</code> yoki <code>-45000 ijara</code>\n"
        "Qo'llanma: /yordam", reply_markup=main_menu(message.from_user.id))


# ═══════════════════════════════════════════════════════ 15. AVTOMATIK HISOBOT
def _admin_ids() -> list[int]:
    return [u["user_id"] for u in users_all() if u["rol"] == "admin"]


def hisobot_soati() -> tuple[int, int]:
    raw = cfg_get("hisobot_vaqti", "22:00")
    try:
        h, m = raw.split(":")
        return int(h), int(m)
    except ValueError:
        return 22, 0


def kun_yakuni_matni(kun: dt.date, biz_ids: list[int]) -> str:
    iso = kun.isoformat()
    lines = [f"🌙 <b>KUN YAKUNI — {kun:%d.%m.%Y}</b>", ""]
    j_inc = j_exp = 0.0
    for b in biz_all():
        if b["id"] not in biz_ids:
            continue
        rows = tx_period(iso, iso, b["id"])
        inc, exp = totals(rows)
        j_inc += inc
        j_exp += exp
        bal = kassa_holati(b["id"])
        lines += [
            f"<b>{b['emoji']} {b['nomi']}</b>",
            f"  🟢 {fmt(inc)}   🔴 {fmt(exp)}",
            f"  {'💰' if inc - exp >= 0 else '⚠️'} Foyda: <b>{fmt(inc - exp)}</b>",
            f"  💵 {fmt(bal['Naqd'])} · 💳 {fmt(bal['Karta'])}",
            "",
        ]
    naqd = sum(kassa_holati(i)["Naqd"] for i in biz_ids)
    karta = sum(kassa_holati(i)["Karta"] for i in biz_ids)
    lines += [
        "═════════════════",
        f"📌 <b>Kunlik foyda: {fmt(j_inc - j_exp)} so'm</b>",
        f"💰 <b>Kassada jami: {fmt(naqd + karta)} so'm</b>",
        f"   💵 {fmt(naqd)} · 💳 {fmt(karta)}",
    ]
    return "\n".join(lines)


def oy_yakuni_matni(kun: dt.date, biz_ids: list[int]) -> str:
    start = kun.replace(day=1).isoformat()
    end = kun.isoformat()
    lines = [f"📅 <b>{MONTHS_UZ[kun.month - 1].upper()} {kun.year} — OY YAKUNI</b>", ""]
    j_inc = j_exp = 0.0
    for b in biz_all():
        if b["id"] not in biz_ids:
            continue
        inc, exp = totals(tx_period(start, end, b["id"]))
        j_inc += inc
        j_exp += exp
        lines += [f"<b>{b['emoji']} {b['nomi']}</b>",
                  f"  🟢 {fmt(inc)}   🔴 {fmt(exp)}",
                  f"  {'💰' if inc - exp >= 0 else '⚠️'} Foyda: <b>{fmt(inc - exp)}</b>", ""]
    rows = [r for r in tx_period(start, end) if r["biznes_id"] in biz_ids]
    by_day: dict[str, float] = defaultdict(float)
    for r in rows:
        by_day[r["sana"]] += r["summa_uzs"] * (1 if r["tur"] == "Kirim" else -1)
    naqd = sum(kassa_holati(i)["Naqd"] for i in biz_ids)
    karta = sum(kassa_holati(i)["Karta"] for i in biz_ids)
    lines += ["═════════════════", f"📌 <b>Oylik foyda: {fmt(j_inc - j_exp)} so'm</b>"]
    if by_day:
        best = max(by_day.items(), key=lambda kv: kv[1])
        lines += [f"🏆 Eng foydali kun: {best[0]} — {fmt(best[1])}",
                  f"📈 O'rtacha kunlik: {fmt(sum(by_day.values()) / len(by_day))}"]
    lines.append(f"💰 Kassa: 💵 {fmt(naqd)} · 💳 {fmt(karta)}")
    return "\n".join(lines)


async def _yubor(uid: int, text: str, fayl_nomi: str, biz_ids: list[int]) -> None:
    try:
        await bot.send_message(uid, text)
        data = await asyncio.to_thread(build_excel, biz_ids)
        await bot.send_document(uid, BufferedInputFile(data, filename=fayl_nomi))
    except Exception as e:                       # bloklangan, chat o'chirilgan va h.k.
        log.warning("Hisobot yuborilmadi (%s): %s", uid, e)


async def hisobot_yubor(kun: dt.date, oylik: bool = False,
                        kechikkan: bool = False) -> None:
    ids = [b["id"] for b in biz_all()]
    if not ids:
        return
    belgi = ("⏳ <i>Kechikkan hisobot — bot o'sha kuni ishlamagan.</i>\n\n"
             if kechikkan else "")
    for uid in _admin_ids():
        await _yubor(uid, belgi + kun_yakuni_matni(kun, ids),
                     f"Hisob-kitob-{kun:%Y-%m-%d}.xlsx", ids)
        if cfg_get("zaxira_yoq", "") != "1":
            await zaxira_yubor(uid, "kunlik")
        if oylik:
            await _yubor(uid, oy_yakuni_matni(kun, ids),
                         f"Hisob-kitob-{MONTHS_UZ[kun.month - 1]}-{kun.year}.xlsx", ids)


def _oyning_oxirgi_kunimi(kun: dt.date) -> bool:
    return (kun + dt.timedelta(days=1)).month != kun.month


async def scheduler() -> None:
    """Belgilangan vaqt kelganda hisobot yuboradi.

    Bot o'chiq bo'lgan kunlar o'tkazib yuborilmaydi — ishga tushganda
    ular ham (kechikkan deb belgilanib) yuboriladi.
    """
    await asyncio.sleep(5)
    while True:
        try:
            now = hozir()
            soat, daqiqa = hisobot_soati()
            bugungi = now.date()
            oxirgi = cfg_get("oxirgi_hisobot", "")

            # o'tkazib yuborilgan kunlar (eng ko'pi 7 kun orqaga)
            if oxirgi:
                try:
                    oxirgi_kun = dt.date.fromisoformat(oxirgi)
                except ValueError:
                    oxirgi_kun = bugungi - dt.timedelta(days=1)
                kun = oxirgi_kun + dt.timedelta(days=1)
                cheklov = bugungi - dt.timedelta(days=7)
                if kun < cheklov:
                    kun = cheklov
                while kun < bugungi:
                    log.info("Kechikkan hisobot: %s", kun)
                    cfg_set("oxirgi_hisobot", kun.isoformat())
                    await hisobot_yubor(kun, oylik=_oyning_oxirgi_kunimi(kun), kechikkan=True)
                    kun += dt.timedelta(days=1)

            await eslatmalar(now)

            if cfg_get("oxirgi_hisobot", "") != bugungi.isoformat() \
                    and (now.hour, now.minute) >= (soat, daqiqa):
                cfg_set("oxirgi_hisobot", bugungi.isoformat())
                log.info("Kunlik hisobot yuborilmoqda...")
                await hisobot_yubor(bugungi, oylik=_oyning_oxirgi_kunimi(bugungi))
        except Exception as e:
            log.warning("Scheduler xatosi: %s", e)
        await asyncio.sleep(30)


async def eslatmalar(now: dt.datetime) -> None:
    """Kunlik eslatmalar: yozuv yo'q · qarz muddati · kassa tekshiruvi."""
    if cfg_get("eslatma_yoq", "") == "1":
        return
    kun = now.date().isoformat()
    try:
        e_soat, e_daqiqa = [int(x) for x in cfg_get("eslatma_vaqti", "20:00").split(":")]
    except ValueError:
        e_soat, e_daqiqa = 20, 0
    if (now.hour, now.minute) < (e_soat, e_daqiqa):
        return
    if cfg_get("oxirgi_eslatma", "") == kun:
        return
    cfg_set("oxirgi_eslatma", kun)

    xabarlar = []

    # 1) bugun yozuv bo'lmagan bizneslar
    bosh = [b for b in biz_all() if not tx_period(kun, kun, b["id"])]
    if bosh:
        royxat = ", ".join(f"{b['emoji']} {b['nomi']}" for b in bosh)
        xabarlar.append(f"📝 Bugun {royxat} bo'yicha yozuv yo'q — "
                        "savdo bo'lmadimi yoki unutildimi?")

    # 2) qarz muddati bugun yoki o'tib ketgan
    kechikkanlar = []
    for d in debts_all(only_open=True):
        if not d["muddat"]:
            continue
        try:
            m = dt.date.fromisoformat(d["muddat"])
        except ValueError:
            continue
        if m <= now.date():
            kun_farq = (now.date() - m).days
            holat = "bugun" if kun_farq == 0 else f"{kun_farq} kun kechikdi"
            kechikkanlar.append(f"   <code>#{d['id']}</code> {d['kim']} — "
                                f"{fmt(d['summa_uzs'] - d['qaytarilgan'])} so'm ({holat})")
    if kechikkanlar:
        xabarlar.append("⏰ <b>Qarz muddati:</b>\n" + "\n".join(kechikkanlar[:10]))

    # 3) kassa uzoq tekshirilmagan
    eski = []
    for b in biz_all():
        oxirgi = cfg_get(f"kassa_tekshiruv_{b['id']}", "")
        if not oxirgi:
            continue
        try:
            farq = (now.date() - dt.date.fromisoformat(oxirgi)).days
        except ValueError:
            continue
        if farq >= 7:
            eski.append(f"   {b['emoji']} {b['nomi']} — {farq} kun oldin")
    if eski:
        xabarlar.append("🧮 <b>Kassa uzoq tekshirilmadi:</b>\n" + "\n".join(eski))

    if not xabarlar:
        return
    matn = "🔔 <b>ESLATMA</b>\n\n" + "\n\n".join(xabarlar) + \
           "\n\n<i>O'chirish: /eslatma yoq</i>"
    for uid in _admin_ids():
        try:
            await bot.send_message(uid, matn)
        except Exception as e:
            log.warning("Eslatma yuborilmadi (%s): %s", uid, e)


@dp.message(Command("eslatma"))
async def cmd_eslatma(message: Message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    hozirgi = cfg_get("eslatma_vaqti", "20:00")
    yoqiq = cfg_get("eslatma_yoq", "") != "1"
    if len(parts) < 2:
        await message.answer(
            f"🔔 <b>Eslatmalar:</b> {'yoqiq ✅' if yoqiq else 'o‘chiq ❌'}\n"
            f"⏰ Vaqti: <b>{hozirgi}</b>\n\n"
            "Bot eslatadi:\n"
            "• bugun yozuv bo'lmasa\n"
            "• qarz muddati kelganda\n"
            "• kassa bir haftadan beri tekshirilmasa\n\n"
            "<code>/eslatma 21:00</code> — vaqtni o'zgartirish\n"
            "<code>/eslatma yoq</code> · <code>/eslatma ha</code>")
        return
    arg = parts[1].strip().lower()
    if arg in ("yoq", "o'chir", "ochir", "off"):
        cfg_set("eslatma_yoq", "1")
        await message.answer("🔕 Eslatmalar o'chirildi.")
        return
    if arg in ("ha", "yoq'", "yoqin", "on"):
        cfg_set("eslatma_yoq", "0")
        await message.answer("🔔 Eslatmalar yoqildi.")
        return
    try:
        h, mn = arg.split(":")
        h, mn = int(h), int(mn)
        assert 0 <= h < 24 and 0 <= mn < 60
    except (ValueError, AssertionError):
        await message.answer("Vaqtni shunday yozing: <code>/eslatma 20:30</code>")
        return
    cfg_set("eslatma_vaqti", f"{h:02d}:{mn:02d}")
    cfg_set("eslatma_yoq", "0")
    await message.answer(f"🔔 Eslatma vaqti: <b>{h:02d}:{mn:02d}</b>")


@dp.message(Command("hisobot"))
async def cmd_hisobot_hozir(message: Message):
    """Hisobotni kutmasdan hozir olish."""
    if not is_admin(message.from_user.id):
        return
    await message.answer("📤 Hisobot tayyorlanmoqda...")
    kun = bugun()
    ids = [b["id"] for b in biz_all()]
    await _yubor(message.from_user.id, kun_yakuni_matni(kun, ids),
                 f"Hisob-kitob-{kun:%Y-%m-%d}.xlsx", ids)


@dp.message(Command("vaqt"))
async def cmd_vaqt(message: Message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) < 2 or ":" not in parts[1]:
        soat, daqiqa = hisobot_soati()
        await message.answer(
            f"Kunlik hisobot vaqti: <b>{soat:02d}:{daqiqa:02d}</b>\n\n"
            "O'zgartirish: <code>/vaqt 21:30</code>")
        return
    try:
        h, m = parts[1].split(":")
        h, m = int(h), int(m)
        assert 0 <= h < 24 and 0 <= m < 60
    except (ValueError, AssertionError):
        await message.answer("Vaqtni shunday yozing: <code>/vaqt 22:00</code>")
        return
    cfg_set("hisobot_vaqti", f"{h:02d}:{m:02d}")
    await message.answer(f"✅ Kunlik hisobot endi soat <b>{h:02d}:{m:02d}</b> da keladi.",
                         reply_markup=main_menu(message.from_user.id))


# ═══════════════════════════════════════════════════════ 15b. ZAXIRA NUSXA
def zaxira_bayt() -> bytes:
    """Bazaning to'liq va xavfsiz nusxasi (yozuv ketayotgan paytda ham)."""
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        nusxa = Path(tmp) / "hisobot.db"
        with db() as manba, sqlite3.connect(nusxa) as maqsad:
            manba.backup(maqsad)
        return nusxa.read_bytes()


async def zaxira_yubor(uid: int, sabab: str = "") -> None:
    try:
        data = await asyncio.to_thread(zaxira_bayt)
        nom = f"zaxira-{hozir():%Y-%m-%d}.db"
        await bot.send_document(
            uid, BufferedInputFile(data, filename=nom),
            caption=f"💾 <b>Zaxira nusxa</b>{(' · ' + sabab) if sabab else ''}\n\n"
                    f"{len(tx_all())} ta yozuv · {len(biz_all())} ta biznes\n"
                    "<i>Bu faylni saqlab qo'ying. Kompyuter buzilsa, uni bot papkasiga "
                    "<code>hisobot.db</code> nomi bilan qo'yasiz — hammasi tiklanadi.</i>")
    except Exception as e:
        log.warning("Zaxira yuborilmadi (%s): %s", uid, e)


@dp.message(Command("zaxira"))
async def cmd_zaxira(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer("💾 Zaxira tayyorlanmoqda...")
    await zaxira_yubor(message.from_user.id, "qo'lda so'ralgan")


# ═══════════════════════════════════════════════════════ 15c. QIDIRUV VA TAHLIL
@dp.message(Command("qidir"))
async def cmd_qidir(message: Message):
    if not allowed(message.from_user.id):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Foydalanish:\n<code>/qidir zapchast</code> — izoh bo'yicha\n"
                             "<code>/qidir 450000</code> — summa bo'yicha")
        return
    soz = parts[1].strip()
    ids = my_ids(message.from_user.id)
    if not ids:
        return
    marks = ",".join("?" * len(ids))
    raqam = parse_amount(soz)
    with db() as c:
        if raqam and soz.replace(" ", "").replace(".", "").replace(",", "").isdigit():
            summa = raqam[0]
            rows = c.execute(
                f"SELECT * FROM amaliyotlar WHERE biznes_id IN ({marks}) "
                "AND ABS(summa_uzs - ?) < 1 ORDER BY id DESC LIMIT 20",
                (*ids, summa)).fetchall()
        else:
            rows = c.execute(
                f"SELECT * FROM amaliyotlar WHERE biznes_id IN ({marks}) "
                "AND LOWER(izoh) LIKE ? ORDER BY id DESC LIMIT 20",
                (*ids, f"%{soz.lower()}%")).fetchall()
    if not rows:
        await message.answer(f"🔍 «{soz}» bo'yicha hech narsa topilmadi.",
                             reply_markup=main_menu(message.from_user.id))
        return
    jami = sum(r["summa_uzs"] * (1 if r["tur"] == "Kirim" else -1) for r in rows)
    lines = [f"🔍 <b>«{soz}»</b> — {len(rows)} ta topildi", ""]
    for r in rows:
        b = biz_get(r["biznes_id"])
        icon = "🟢" if r["tur"] == "Kirim" else "🔴"
        lines.append(f"{icon} <code>#{r['id']}</code> {b['emoji'] if b else ''} {r['sana']} · "
                     f"<b>{fmt(r['summa_uzs'])}</b>" + (f" · {r['izoh']}" if r["izoh"] else ""))
    lines.append(f"\n<b>Jami: {fmt(jami)} so'm</b>")
    lines.append("<i>Tahrirlash: /tahrir RAQAM</i>")
    await message.answer("\n".join(lines), reply_markup=main_menu(message.from_user.id))


@dp.message(Command("kunlar"))
async def cmd_kunlar(message: Message):
    """Hafta kunlari bo'yicha o'rtacha — qaysi kun kuchli."""
    if not allowed(message.from_user.id):
        return
    ids = my_ids(message.from_user.id)
    start = (bugun() - dt.timedelta(days=29)).isoformat()
    end = bugun().isoformat()
    NOM = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
    lines = ["📅 <b>HAFTA KUNLARI</b>", "<i>oxirgi 30 kun · o'rtacha kunlik foyda</i>", ""]
    for b in biz_all():
        if b["id"] not in ids:
            continue
        yigindi = {i: [0.0, set()] for i in range(7)}
        for r in tx_period(start, end, b["id"]):
            kun = dt.date.fromisoformat(r["sana"])
            k = kun.weekday()
            yigindi[k][0] += r["summa_uzs"] * (1 if r["tur"] == "Kirim" else -1)
            yigindi[k][1].add(r["sana"])
        ortacha = {i: (v[0] / len(v[1]) if v[1] else 0) for i, v in yigindi.items()}
        eng = max(ortacha.values()) or 1
        lines.append(f"<b>{b['emoji']} {b['nomi']}</b>")
        for i in sorted(ortacha, key=lambda x: -ortacha[x]):
            if not yigindi[i][1]:
                continue
            uzunlik = max(1, round(ortacha[i] / eng * 8)) if ortacha[i] > 0 else 0
            lines.append(f"  {NOM[i][:3]} {'█' * uzunlik:<8} {fmt(ortacha[i])}")
        lines.append("")
    await message.answer("\n".join(lines).strip(), reply_markup=main_menu(message.from_user.id))


@dp.message(Command("xarajat"))
async def cmd_xarajat(message: Message):
    """Izohlar bo'yicha guruhlangan xarajatlar."""
    if not allowed(message.from_user.id):
        return
    ids = my_ids(message.from_user.id)
    t = bugun()
    start, end = t.replace(day=1).isoformat(), t.isoformat()
    lines = [f"🔴 <b>{MONTHS_UZ[t.month - 1].upper()} · XARAJATLAR</b>", ""]
    bor = False
    for b in biz_all():
        if b["id"] not in ids:
            continue
        guruh: dict[str, list] = {}
        for r in tx_period(start, end, b["id"]):
            if r["tur"] != "Chiqim":
                continue
            kalit = (r["izoh"] or "izohsiz").strip().lower()[:28]
            g = guruh.setdefault(kalit, [0.0, 0])
            g[0] += r["summa_uzs"]
            g[1] += 1
        if not guruh:
            continue
        bor = True
        jami = sum(v[0] for v in guruh.values())
        lines.append(f"<b>{b['emoji']} {b['nomi']}</b> — jami {fmt(jami)}")
        for k, v in sorted(guruh.items(), key=lambda kv: -kv[1][0])[:8]:
            ulush = v[0] / jami * 100
            lines.append(f"  {k} — <b>{fmt(v[0])}</b> "
                         f"<i>({v[1]} marta · {ulush:.0f}%)</i>")
        lines.append("")
    if not bor:
        lines.append("Bu oyda chiqim yozilmagan.")
    await message.answer("\n".join(lines).strip(), reply_markup=main_menu(message.from_user.id))


# ═══════════════════════════════════════════════════════ 16. MINI APP
def webapp_url() -> str:
    """https manzil sozlangan bo'lsa Mini App tugmasi ishlaydi."""
    return cfg_get("webapp_url", "").strip()


@dp.message(Command("ilova", "app"))
async def cmd_ilova(message: Message):
    if not allowed(message.from_user.id):
        return
    url = webapp_url()
    if url.startswith("https://"):
        await message.answer(
            "📱 <b>Hisob-kitob ilovasi</b>\n\n"
            "Telegramdan chiqmasdan ochiladi: kassa, yozish, tarix, hisobot.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="📱 Ilovani ochish", web_app=WebAppInfo(url=url))
            ]]))
    else:
        await message.answer(
            "📱 <b>Mini App hozircha faqat kompyuterda sinov rejimida.</b>\n\n"
            f"Brauzerda oching:\n<code>http://localhost:{WEB_PORT}/?dev=1</code>\n\n"
            "Telegram ichida ochilishi uchun bot internetdagi serverda (https) "
            "turishi kerak. Server tayyor bo'lgach:\n"
            "<code>/ilova_manzil https://sizning-manzilingiz</code>")


@dp.message(Command("ilova_manzil"))
async def cmd_ilova_manzil(message: Message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip().startswith("https://"):
        await message.answer("Foydalanish: <code>/ilova_manzil https://mysite.com</code>")
        return
    cfg_set("webapp_url", parts[1].strip().rstrip("/"))
    await message.answer("✅ Manzil saqlandi. Endi /ilova tugmasi Telegram ichida ochiladi.")


def _bulut_domenini_yoz() -> None:
    """Railway/Render bergan domenni Mini App manzili sifatida saqlaydi."""
    for kalit in ("RAILWAY_PUBLIC_DOMAIN", "RENDER_EXTERNAL_HOSTNAME"):
        d = os.getenv(kalit, "").strip()
        if not d:
            continue
        d = d.replace("https://", "").replace("http://", "").rstrip("/")
        url = "https://" + d
        if cfg_get("webapp_url", "").strip() != url:
            cfg_set("webapp_url", url)
            print(f"[mini app] Manzil avtomatik saqlandi: {url}")
        return


async def main():
    init_db()
    _bulut_domenini_yoz()
    me = await bot.get_me()
    nomlar = ", ".join(b["nomi"] for b in biz_all()) or "(hali yo'q)"
    print("\n" + "═" * 58)
    print(f"  ✅ Bot ishga tushdi:  @{me.username}")
    print(f"  📂 Baza:              {DB_PATH}")
    print(f"  🏢 Bizneslar:         {nomlar}")
    soat, daqiqa = hisobot_soati()
    print(f"  🔗 Havola:            https://t.me/{me.username}")
    print(f"  🌙 Kunlik hisobot:    har kuni {soat:02d}:{daqiqa:02d}")
    try:
        local = webapp.start(sys.modules[__name__], TOKEN, WEB_PORT,
                             dev=WEB_DEV, bind=WEB_HOST)
        rejim = "sinov rejimi" if WEB_DEV else "faqat Telegram imzosi bilan"
        print(f"  📱 Mini App:          {local}  ({rejim})")
    except OSError as e:
        print(f"  ⚠️  Mini App ishga tushmadi ({e}). Port band bo'lishi mumkin.")
    if not has_admin():
        print("\n  👉 Telegramda botga /start yuboring — birinchi kirgan")
        print("     odam avtomatik EGA (admin) bo'ladi.")
    print("\n  To'xtatish uchun: Ctrl+C")
    print("═" * 58 + "\n")
    await bot.delete_webhook(drop_pending_updates=True)
    asyncio.create_task(scheduler())
    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        print("\nBot to'xtatildi.")